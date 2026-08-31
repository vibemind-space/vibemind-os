"""
Introspection Router -- FastAPI endpoints for brain self-monitoring (Default Mode Network).

Extracted from the old brain_dashboard_server.py (~30+ routes).  Covers:
  - Brain state / gates / activation / strategies
  - Emotional / homeostatic / memory state (graceful fallback)
  - Cognitive loop / agent loop state (graceful fallback)
  - Health sub-endpoints (components, dependencies, readiness, liveness)
  - Frequency controller
  - Monitoring / observability (metrics, audit trail, loop traces, error rates, heatmap)
  - Goals / evolution / CTM / cognitive status
  - Causal / meta / federated subsystems
  - LLM stats
  - Heartbeat / consciousness / neuromodulation proxies
  - Conversation monitoring / simulation
  - Advanced learning health
  - Sensory extract / predict path

All state lives on ``request.app.state.<module>`` attributes.
Every route gracefully handles ``module is None`` (testing mode) —
no HTTP proxying to localhost:5003.
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
import time
from pathlib import Path
from typing import Any, Dict, Optional

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse

logger = logging.getLogger(__name__)
router = APIRouter()


def _trace_lite(pe, trace_id: str, intent: str, routed_via: str, final_text: str = "") -> None:
    """E2E-Trace (2026-06-09): plan-lose Zweige (meta/easy/som/som-team/no-plan) in
    den Trace schreiben, damit GET /api/trace/{id} JEDE Anfrage zeigt — nicht nur
    die mit PlanExecutor-Plan. Best-effort, nie blockierend/500."""
    try:
        rec = getattr(pe, "recorder", None)
        if rec is not None:
            rec.record_lite(trace_id, intent, routed_via, final_text)
    except Exception as e:  # noqa: BLE001
        logger.debug(f"[trace] record_lite skipped: {e}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def convert_numpy(obj: Any) -> Any:
    """Convert NumPy types to native Python types for JSON serialisation."""
    try:
        import numpy as np
    except ImportError:
        return obj

    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        return float(obj)
    elif isinstance(obj, (np.bool_,)):
        return bool(obj)
    elif isinstance(obj, dict):
        return {k: convert_numpy(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [convert_numpy(i) for i in obj]
    return obj



# ===================================================================
# Group 1 — Brain State  (meta_router, brain_monitor, strategy_lib)
# ===================================================================

@router.get("/api/brain/state")
async def brain_state(request: Request):
    """Current brain routing state."""
    mr = request.app.state.meta_router
    if mr is None:
        return JSONResponse({
            "state": None,
            "message": "meta_router not initialized",
            "timestamp": time.time(),
        })
    try:
        state = mr.get_state()
        return JSONResponse({
            "state": convert_numpy(state),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "state": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/brain/gates")
async def brain_gates(request: Request):
    """Current gate values from BrainActivityMonitor."""
    bm = request.app.state.brain_monitor
    if bm is None:
        return JSONResponse({
            "gates": None,
            "message": "brain_monitor not initialized",
            "timestamp": time.time(),
        })
    try:
        gates = bm.get_gates()
        return JSONResponse({
            "gates": convert_numpy(gates),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "gates": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/brain/activation")
async def brain_activation(request: Request):
    """Activation map from BrainActivityMonitor."""
    bm = request.app.state.brain_monitor
    if bm is None:
        return JSONResponse({
            "activation": None,
            "message": "brain_monitor not initialized",
            "timestamp": time.time(),
        })
    try:
        activation = bm.get_activation()
        return JSONResponse({
            "activation": convert_numpy(activation),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "activation": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/brain/strategies")
async def brain_strategies(request: Request):
    """Available strategies from StrategyLibrary."""
    sl = request.app.state.strategy_lib
    if sl is None:
        return JSONResponse({
            "strategies": None,
            "message": "strategy_lib not initialized",
            "timestamp": time.time(),
        })
    try:
        strategies = sl.list_strategies()
        return JSONResponse({
            "strategies": convert_numpy(strategies),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "strategies": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/brain/interventions")
async def brain_interventions(request: Request):
    """Recent interventions from LiveBrainMonitor."""
    lm = request.app.state.live_monitor
    if lm is None:
        return JSONResponse({
            "interventions": [],
            "message": "live_monitor not initialized",
            "timestamp": time.time(),
        })
    try:
        interventions = lm.get_interventions()
        return JSONResponse({
            "interventions": convert_numpy(interventions),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "interventions": [],
            "error": str(exc),
            "timestamp": time.time(),
        })


# ===================================================================
# Group 2 — Proxy-replacement routes (graceful fallbacks)
# ===================================================================

@router.get("/api/brain/cognitive_loop")
async def cognitive_loop(request: Request):
    """Cognitive loop state — graceful fallback when unified brain is not wired."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "cognitive loop not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/agent_loop_state")
async def agent_loop_state(request: Request):
    """Agent loop state — reads live AgentLoop if wired, else graceful fallback."""
    al = getattr(request.app.state, "agent_loop", None)
    if al is None:
        return JSONResponse({
            "enabled": False,
            "state": None,
            "message": "agent loop not connected to unified brain",
            "timestamp": time.time(),
        })
    try:
        get_state = getattr(al, "get_state", None) or getattr(al, "state", None)
        if callable(get_state):
            state_snapshot = get_state()
        else:
            state_snapshot = {
                "state": getattr(al, "_state", None) or str(getattr(al, "state", "unknown")),
                "tick_count": getattr(al, "_tick_count", None),
                "has_radial": getattr(al, "radial_network", None) is not None,
                "has_seed_encoder": getattr(al, "seed_encoder", None) is not None,
                "has_experience_buffer": getattr(al, "experience_buffer", None) is not None,
            }
        return JSONResponse({
            "enabled": True,
            "state": convert_numpy(state_snapshot),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "enabled": True,
            "state": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.post("/api/brain/agent_loop/submit")
async def agent_loop_submit(request: Request):
    """Submit task to agent loop — 503 until unified brain is wired."""
    return JSONResponse(
        {"error": "agent loop not initialized", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/brain/emotional_state")
async def emotional_state(request: Request):
    """Emotional state — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "emotional state not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/homeostatic_state")
async def homeostatic_state(request: Request):
    """Homeostatic state — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "homeostatic state not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/memory_state")
async def memory_state(request: Request):
    """Memory state — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "memory state not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/heartbeat_status")
async def heartbeat_status(request: Request):
    """Heartbeat status — graceful fallback."""
    return JSONResponse({
        "active": False,
        "message": "heartbeat not connected to unified brain",
        "timestamp": time.time(),
    })


@router.post("/api/brain/sensory_extract")
async def sensory_extract(request: Request):
    """Sensory extract — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "message": "sensory extract not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/goal_graph_state")
async def goal_graph_state(request: Request):
    """Goal graph state — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "goal graph not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/neuromodulation_state")
async def neuromodulation_state(request: Request):
    """Neuromodulation state — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "neuromodulation not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/consciousness_state")
async def consciousness_state(request: Request):
    """Consciousness state — graceful fallback."""
    return JSONResponse({
        "enabled": False,
        "state": None,
        "message": "consciousness module not connected to unified brain",
        "timestamp": time.time(),
    })


# ===================================================================
# Group 3 — Monitoring & Observability
# ===================================================================

@router.get("/api/brain/metrics")
async def brain_metrics(request: Request):
    """Prometheus-style metrics — returns plain text."""
    return PlainTextResponse(
        "# Metrics unavailable — unified brain not connected\n",
        media_type="text/plain",
    )


@router.get("/api/brain/metrics_json")
async def brain_metrics_json(request: Request):
    """JSON metrics — graceful fallback."""
    return JSONResponse({
        "error": "metrics unavailable",
        "timestamp": time.time(),
    })


@router.get("/api/brain/audit_trail")
async def audit_trail(request: Request):
    """Recent audit trail entries."""
    return JSONResponse({
        "recent": [],
        "stats": {},
        "timestamp": time.time(),
    })


@router.get("/api/brain/loop_traces")
async def loop_traces(request: Request):
    """Cognitive loop traces."""
    return JSONResponse({
        "recent_traces": [],
        "phase_stats": {},
        "total_traces": 0,
        "timestamp": time.time(),
    })


@router.get("/api/brain/error_rates")
async def error_rates(request: Request):
    """Error rates by component."""
    return JSONResponse({
        "error_rates": {},
        "recent_errors": [],
        "timestamp": time.time(),
    })


@router.get("/api/brain/heatmap")
async def heatmap(request: Request):
    """Modality activation heatmap data."""
    return JSONResponse({
        "heatmap": {
            "modalities": [],
            "matrix": [],
        },
        "modality_averages": {},
        "timestamp": time.time(),
    })


# ===================================================================
# Group 4 — Frequency Controller
# ===================================================================

@router.get("/api/brain/frequency")
async def brain_frequency(request: Request):
    """Current brain frequency state."""
    fc = request.app.state.frequency_controller
    if fc is None:
        return JSONResponse({
            "frequency": None,
            "message": "frequency_controller not initialized",
            "timestamp": time.time(),
        })
    try:
        state = fc.get_state()
        return JSONResponse({
            "frequency": convert_numpy(state),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "frequency": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.post("/api/brain/frequency/set")
async def brain_frequency_set(request: Request):
    """Set brain frequency parameters."""
    fc = request.app.state.frequency_controller
    if fc is None:
        return JSONResponse({
            "ok": False,
            "message": "frequency_controller not initialized",
            "timestamp": time.time(),
        })
    try:
        body = await request.json()
        # Whitelist accepted keys to avoid parameter injection
        allowed = {"mode", "band", "target_hz", "ramp_time", "activation", "suppress_others"}
        filtered = {k: v for k, v in body.items() if k in allowed}
        result = fc.set_frequency(**filtered)
        return JSONResponse({
            "ok": True,
            "result": convert_numpy(result),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "ok": False,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/brain/frequency/bands")
async def brain_frequency_bands(request: Request):
    """Available frequency bands."""
    fc = request.app.state.frequency_controller
    if fc is None:
        return JSONResponse({
            "bands": None,
            "message": "frequency_controller not initialized",
            "timestamp": time.time(),
        })
    try:
        bands = fc.get_bands()
        return JSONResponse({
            "bands": convert_numpy(bands),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "bands": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/brain/frequency/markers")
async def brain_frequency_markers(request: Request):
    """Frequency event markers."""
    fc = request.app.state.frequency_controller
    if fc is None:
        return JSONResponse({
            "markers": None,
            "message": "frequency_controller not initialized",
            "timestamp": time.time(),
        })
    try:
        markers = fc.get_markers()
        return JSONResponse({
            "markers": convert_numpy(markers),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "markers": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


# ===================================================================
# Group 5 — Health (system-level)
# ===================================================================

@router.get("/api/health/space-registry")
def space_registry_health():
    """Canonical space/event/executor catalog consistency (no live fallback)."""
    import yaml

    from core.space_contract import registry_health

    capabilities_path = Path(__file__).resolve().parents[2] / "data" / "capabilities.yaml"
    capabilities = yaml.safe_load(capabilities_path.read_text(encoding="utf-8")) or []
    return JSONResponse(registry_health(capabilities=capabilities))

@router.get("/api/health/components")
async def health_components(request: Request):
    """Health of individual brain components."""
    state = request.app.state
    components = {
        "meta_router": state.meta_router is not None,
        "brain_monitor": state.brain_monitor is not None,
        "strategy_lib": state.strategy_lib is not None,
        "live_monitor": state.live_monitor is not None,
        "path_planner": state.path_planner is not None,
        "llm_router": state.llm_router is not None,
        "frequency_controller": state.frequency_controller is not None,
        "oscillator": state.oscillator is not None,
        "checkpoint_manager": state.checkpoint_manager is not None,
        "swarm_orchestrator": state.swarm_orchestrator is not None,
    }
    total = len(components)
    healthy = sum(1 for v in components.values() if v)
    return JSONResponse({
        "components": components,
        "healthy": healthy,
        "total": total,
        "status": "healthy" if healthy == total else ("degraded" if healthy > 0 else "not_initialized"),
        "timestamp": time.time(),
    })


@router.get("/api/health/dependencies")
async def health_dependencies(request: Request):
    """External dependency health."""
    return JSONResponse({
        "dependencies": {
            "unified_brain": False,
            "memory_api": False,
            "llm_service": False,
        },
        "message": "dependency checks not yet wired",
        "timestamp": time.time(),
    })


@router.get("/api/health/readiness")
async def health_readiness(request: Request):
    """Kubernetes-style readiness probe."""
    return JSONResponse({
        "ready": True,
        "timestamp": time.time(),
    })


@router.get("/api/health/liveness")
async def health_liveness(request: Request):
    """Kubernetes-style liveness probe."""
    return JSONResponse({
        "alive": True,
        "timestamp": time.time(),
    })


# ===================================================================
# Group 6 — LLM Stats
# ===================================================================

@router.get("/api/llm/probe")
async def llm_probe(request: Request):
    """Diagnostic probe: is MicroAgentPool wired + can it call LLM?"""
    state = request.app.state
    pool = getattr(state, "micro_agent_pool", None)
    init_error = getattr(state, "micro_agent_pool_error", None)
    bc = getattr(state, "brain_chat", None)
    bc_pool = getattr(bc, "_micro_agent_pool", None) if bc else None

    result = {
        "state_pool_present": pool is not None,
        "brain_chat_pool_present": bc_pool is not None,
        "pool_same_instance": (pool is bc_pool) if (pool and bc_pool) else False,
        "init_error": init_error,
    }
    if pool is not None:
        result["pool_router_present"] = pool._router is not None
        result["pool_agents"] = list(pool._agents.keys()) if hasattr(pool, "_agents") else []
        result["pool_total_runs"] = getattr(pool, "_total_runs", None)
        result["pool_total_failures"] = getattr(pool, "_total_failures", None)
        # Try a live call with the responder
        try:
            txt = pool._call_agent("responder", "Say the word PONG and nothing else.")
            result["live_responder_call"] = txt[:200] if txt else None
            result["live_ok"] = bool(txt)
        except Exception as exc:
            result["live_responder_call"] = None
            result["live_ok"] = False
            result["live_error"] = str(exc)
    return JSONResponse(result)


@router.get("/api/llm/stats")
async def llm_stats(request: Request):
    """LLM routing statistics."""
    lr = request.app.state.llm_router
    if lr is None:
        return JSONResponse(
            {"error": "llm_router not initialized", "timestamp": time.time()},
            status_code=503,
        )
    try:
        stats = lr.get_statistics()
        # Phase 11.T.5 — surface concurrency state alongside call counts.
        try:
            from core.multi_llm_router import get_llm_concurrency_stats
            stats["concurrency"] = get_llm_concurrency_stats()
        except Exception:
            pass
        return JSONResponse({
            "stats": convert_numpy(stats),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "stats": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


# ===================================================================
# Group 6b — Knowledge Graph (Qdrant-backed unified KG)
# ===================================================================

@router.get("/api/kg/stats")
async def kg_stats(request: Request):
    """Stats about the unified knowledge graph."""
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse(
            {"error": "qdrant_kg not initialized", "timestamp": time.time()},
            status_code=503,
        )
    try:
        from core.qdrant_kg import COLLECTIONS
        per_coll: Dict[str, Any] = {}
        total = 0
        for logical, name in COLLECTIONS.items():
            try:
                info = kg.client.get_collection(name)
                per_coll[logical] = {
                    "qdrant_name": name,
                    "points_count": info.points_count,
                }
                total += info.points_count
            except Exception as e:
                per_coll[logical] = {
                    "qdrant_name": name,
                    "points_count": 0,
                    "error": str(e),
                }
        return JSONResponse({
            "collections": per_coll,
            "total_points": total,
            "stats": convert_numpy(dict(kg.stats)),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.get("/api/kg/search")
async def kg_search(
    request: Request, q: str, node_type: str = "", collection: str = "",
    limit: int = 10, threshold: float = 0.0,
):
    """Semantic kNN search across cognitive collections.

    Query params:
        q: text query (multilingual via Qwen)
        node_type: optional filter (thought/response/bubble/idea/space/event/snapshot)
        collection: optional logical collection name
            (episodic|semantic|procedural|state|artifacts). If empty,
            searches all cognitive collections and merges by score.
        limit: max hits
        threshold: min cosine score
    """
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse(
            {"error": "qdrant_kg not initialized", "timestamp": time.time()},
            status_code=503,
        )
    try:
        nt = node_type or None
        coll = collection or None
        # Phase 11.U.B — kg.search hits a sync embedder + sync Qdrant client.
        # Run in a thread so the FastAPI event loop stays responsive.
        hits = await asyncio.to_thread(
            kg.search, q, node_type=nt, collection=coll,
            limit=int(limit), score_threshold=float(threshold),
        )
        return JSONResponse({
            "query": q, "node_type": nt, "collection": coll,
            "count": len(hits),
            "hits": convert_numpy(hits), "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.get("/api/kg/route")
async def kg_route(
    request: Request, q: str, limit: int = 3, threshold: float = 0.3,
):
    """Replace space_routing_head.pt / event_routing_head.pt with graph
    kNN search.

    Returns top-k spaces AND top-k events with scores. Brain can blend
    these into its Thalamus routing priors. Every successful route can
    later bump activation_strength on the chosen space/event point,
    yielding usage-weighted routing without gradient descent.
    """
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse(
            {"error": "qdrant_kg not initialized", "timestamp": time.time()},
            status_code=503,
        )
    try:
        # Phase 11.U.B — both kg.search calls offloaded to threads. They run
        # in parallel via asyncio.gather to keep total latency unchanged.
        spaces, events = await asyncio.gather(
            asyncio.to_thread(
                kg.search, q, node_type="space", limit=int(limit),
                score_threshold=float(threshold),
            ),
            asyncio.to_thread(
                kg.search, q, node_type="event", limit=int(limit),
                score_threshold=float(threshold),
            ),
        )
        # Normalize: show only id, score, title for a clean routing payload
        def _trim(hits, id_key):
            return [{
                "id": h["payload"].get(id_key) or h["id"],
                "score": h["score"],
                "title": h["payload"].get("title", ""),
                "target_space": h["payload"].get("target_space"),
            } for h in hits]
        return JSONResponse({
            "query": q,
            "spaces": _trim(spaces, "space_id"),
            "events": _trim(events, "event_id"),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.post("/api/kg/confirm_route")
async def kg_confirm_route(request: Request):
    """Bump activation_strength on a chosen space/event after a route
    actually worked. Body: {"kind": "space|event", "id": "<external_id>",
    "delta": 1.0}. No gradient descent — just usage-weighted priors that
    rise over time and make future kNN searches gravitate toward
    historically successful choices.
    """
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse(
            {"error": "qdrant_kg not initialized"}, status_code=503,
        )
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    kind = body.get("kind") or ""
    ext_id = body.get("id") or ""
    delta = float(body.get("delta", 1.0))
    if kind not in ("space", "event") or not ext_id:
        return JSONResponse({
            "error": "need kind=space|event and id",
        }, status_code=400)
    try:
        import hashlib, uuid as _uuid
        # Spaces + events both live in the procedural collection.
        coll_name = kg.collection_for(kind)
        pid = str(_uuid.UUID(
            hashlib.sha256(ext_id.encode("utf-8")).hexdigest()[:32]
        ))
        rec = kg.client.retrieve(
            collection_name=coll_name, ids=[pid], with_payload=True,
        )
        if not rec:
            return JSONResponse({
                "error": f"no point for {kind}:{ext_id} in {coll_name}",
            }, status_code=404)
        payload = rec[0].payload or {}
        current = float(payload.get("activation_strength", 0.0) or 0.0)
        new_val = current + delta
        kg.client.set_payload(
            collection_name=coll_name,
            payload={"activation_strength": new_val},
            points=[pid],
        )
        return JSONResponse({
            "kind": kind, "id": ext_id, "collection": coll_name,
            "prev": current, "new": new_val, "delta": delta,
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.get("/api/kg/related")
async def kg_related(request: Request, point_id: str):
    """Return the linked.* edges of a point by its Qdrant UUID."""
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse(
            {"error": "qdrant_kg not initialized", "timestamp": time.time()},
            status_code=503,
        )
    try:
        from core.qdrant_kg import COLLECTIONS
        # Scan all cognitive collections until we find the point.
        rec = None
        found_coll = None
        for logical, name in COLLECTIONS.items():
            try:
                r = kg.client.retrieve(
                    collection_name=name, ids=[point_id], with_payload=True,
                )
                if r:
                    rec = r
                    found_coll = logical
                    break
            except Exception:
                continue
        if not rec:
            return JSONResponse({
                "point_id": point_id, "found": False,
                "timestamp": time.time(),
            })
        payload = rec[0].payload or {}
        return JSONResponse({
            "collection": found_coll,
            "point_id": point_id, "found": True,
            "node_type": payload.get("node_type"),
            "content": (payload.get("content") or "")[:500],
            "linked": payload.get("linked") or {},
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.post("/api/brain/subagent")
async def brain_subagent(request: Request):
    """Dispatch a focused subtask to an LLM subagent (Claude or Groq).

    Body: {
      "tool": "claude_subagent" | "groq_subagent" (default: claude_subagent)
      "prompt": "<the actual task>",
      "system": "<optional system prompt>",
      "model": "<override model id>",
      "max_tokens": <int>,
      "temperature": <float>
    }

    Returns: {ok, tool, model, text, latency_ms, error?}
    """
    disp = getattr(request.app.state, "subagent_dispatcher", None)
    if disp is None:
        return JSONResponse(
            {"error": "subagent_dispatcher not available"}, status_code=503,
        )
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    tool = body.get("tool") or "claude_subagent"
    prompt = body.get("prompt") or ""
    if not prompt:
        return JSONResponse({"error": "prompt required"}, status_code=400)
    kwargs = {k: v for k, v in body.items() if k not in ("tool",)}
    try:
        result = disp.dispatch(tool, **kwargs)
        return JSONResponse(result)
    except Exception as exc:
        return JSONResponse({
            "ok": False, "tool": tool, "error": str(exc),
        }, status_code=500)


@router.get("/api/brain/subagent/stats")
async def brain_subagent_stats(request: Request):
    """Stats about subagent dispatcher (calls per tool, failures, etc.)."""
    disp = getattr(request.app.state, "subagent_dispatcher", None)
    if disp is None:
        return JSONResponse(
            {"enabled": False, "message": "subagent_dispatcher not running"},
        )
    return JSONResponse({
        "enabled": True,
        "stats": convert_numpy(dict(disp.stats)),
        "timestamp": time.time(),
    })


@router.post("/api/brain/dispatch")
async def brain_dispatch(request: Request):
    """Dispatch a task to one or more Minibook agents.

    Body: {
      "project_id": "<uuid, optional - default VibeMind Collaboration>",
      "agents": ["vibemind_ideas", ...],
      "intent": "what should they do",
      "task_spec": {...optional structured payload...}
    }

    Returns: {post_id, project_id, agents, online, ts}
    Brain can later poll comments via the same minibook_client.
    """
    DEFAULT_PROJECT = "46daa2f8-6f39-4cde-87c3-c95235bfb557"  # VibeMind Collaboration
    mb = getattr(request.app.state, "minibook_client", None)
    if mb is None:
        al = getattr(request.app.state, "agent_loop", None)
        mb = getattr(al, "minibook_client", None) if al is not None else None
    if mb is None:
        return JSONResponse(
            {"error": "minibook_client not available"}, status_code=503,
        )
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    project_id = body.get("project_id") or DEFAULT_PROJECT
    agents = body.get("agents") or []
    intent = body.get("intent") or ""
    task_spec = body.get("task_spec")
    if not agents or not intent:
        return JSONResponse({
            "error": "need agents (list) and intent (str)",
        }, status_code=400)
    try:
        post_id = mb.dispatch_task(
            project_id=project_id,
            target_agents=agents,
            intent=intent,
            task_spec=task_spec,
        )
        return JSONResponse({
            "post_id": post_id,
            "project_id": project_id,
            "agents": agents,
            "online": getattr(mb, "is_online", None),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.get("/api/brain/dispatch/{post_id}/comments")
async def brain_dispatch_comments(request: Request, post_id: str):
    """Retrieve agent replies on a previously dispatched task."""
    mb = getattr(request.app.state, "minibook_client", None)
    if mb is None:
        al = getattr(request.app.state, "agent_loop", None)
        mb = getattr(al, "minibook_client", None) if al is not None else None
    if mb is None:
        return JSONResponse(
            {"error": "minibook_client not available"}, status_code=503,
        )
    try:
        comments = mb.get_comments(post_id)
        return JSONResponse({
            "post_id": post_id,
            "count": len(comments),
            "comments": comments,
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.post("/api/kg/consolidate")
async def kg_consolidate_now(request: Request):
    """Trigger one consolidation pass (episodic -> semantic). Returns summary."""
    ce = getattr(request.app.state, "consolidation_engine", None)
    if ce is None:
        return JSONResponse(
            {"error": "consolidation_engine not running"}, status_code=503,
        )
    try:
        summary = ce.run_once()
        return JSONResponse({
            "ok": True,
            "summary": convert_numpy(summary),
            "stats": convert_numpy(dict(ce.stats)),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "ok": False, "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.get("/api/kg/consolidation_stats")
async def kg_consolidation_stats(request: Request):
    """Stats about the consolidation engine."""
    ce = getattr(request.app.state, "consolidation_engine", None)
    if ce is None:
        return JSONResponse(
            {"enabled": False, "message": "consolidation_engine not running"},
        )
    return JSONResponse({
        "enabled": True,
        "stats": convert_numpy(dict(ce.stats)),
        "timestamp": time.time(),
    })


@router.post("/api/kg/snapshot")
async def kg_snapshot_now(request: Request):
    """Capture one Brain self-state snapshot into brain-state. Returns snapshot_id."""
    se = getattr(request.app.state, "snapshot_engine", None)
    if se is None:
        return JSONResponse(
            {"error": "snapshot_engine not running"}, status_code=503,
        )
    try:
        sid = se.snapshot_now()
        return JSONResponse({
            "ok": sid is not None,
            "snapshot_id": sid,
            "stats": convert_numpy(dict(se.stats)),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "ok": False, "error": str(exc), "timestamp": time.time(),
        }, status_code=500)


@router.get("/api/kg/snapshot_stats")
async def kg_snapshot_stats(request: Request):
    """Stats about the snapshot engine."""
    se = getattr(request.app.state, "snapshot_engine", None)
    if se is None:
        return JSONResponse(
            {"enabled": False, "message": "snapshot_engine not running"},
        )
    return JSONResponse({
        "enabled": True,
        "stats": convert_numpy(dict(se.stats)),
        "timestamp": time.time(),
    })


@router.get("/api/kg/snapshots")
async def kg_list_snapshots(request: Request, limit: int = 20):
    """List recent snapshots (newest first), payload-only (no vectors)."""
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse({"error": "qdrant_kg not available"}, status_code=503)
    try:
        from core.qdrant_kg import COLLECTIONS
        from qdrant_client.http import models as qm
        coll = COLLECTIONS["state"]
        limit = max(1, min(200, int(limit)))
        # Scroll, then sort by created_at desc client-side
        batch, _ = kg.client.scroll(
            collection_name=coll, limit=500, with_payload=True, with_vectors=False,
        )
        rows = []
        for rec in batch:
            p = rec.payload or {}
            if p.get("node_type") != "snapshot":
                continue
            rows.append({
                "point_id": str(rec.id),
                "snapshot_id": p.get("snapshot_id"),
                "ts": p.get("ts"),
                "created_at": p.get("created_at"),
                "content": p.get("content", "")[:300],
                "modulation": p.get("modulation"),
                "state_summary": p.get("state_summary"),
            })
        rows.sort(key=lambda r: r.get("ts") or 0, reverse=True)
        return JSONResponse({
            "count": len(rows[:limit]),
            "total": len(rows),
            "snapshots": rows[:limit],
        })
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.post("/api/tribe/predict")
async def tribe_predict(request: Request):
    """TriBE v2 text -> cortical activation -> bridge levels.

    Body: {"text": "<some text>"}
    Returns: {ok, text, rois, bridges, shape, latency_ms}

    If TriBE weights aren't available (gated Llama-3.2-3B still
    awaiting Meta approval) and TRIBE_DUMMY=1, a deterministic
    pseudo-vector is used so downstream wiring stays testable.
    """
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    text = (body.get("text") or "").strip()
    if not text:
        return JSONResponse({"error": "need 'text'"}, status_code=400)
    try:
        from core.tribe_encoder import TribeEncoder
        enc = TribeEncoder.get()
        t0 = time.time()
        vec = enc.predict(text)
        dt = (time.time() - t0) * 1000
        if vec is None:
            return JSONResponse({
                "ok": False,
                "error": enc.stats.get("last_error")
                         or enc._load_error
                         or "predict returned None",
                "status": enc.status(),
            }, status_code=503)
        rois = enc.aggregate_roi(vec)
        bridges = enc.bridge_levels(vec)
        return JSONResponse({
            "ok": True,
            "text": text[:500],
            "shape": [int(vec.shape[0])],
            "rois": {k: round(float(v), 6) for k, v in rois.items()},
            "bridges": {k: round(float(v), 6) for k, v in bridges.items()},
            "latency_ms": round(dt, 2),
            "mode": enc.stats.get("last_mode", "real"),
        })
    except Exception as exc:
        return JSONResponse({
            "ok": False, "error": str(exc),
        }, status_code=500)


@router.get("/api/tribe/status")
async def tribe_status(request: Request):
    """Diagnostic: is TriBE loaded / what's the last error / stats."""
    try:
        from core.tribe_encoder import TribeEncoder
        enc = TribeEncoder.get()
        return JSONResponse(enc.status())
    except Exception as exc:
        return JSONResponse({"enabled": False, "error": str(exc)})


@router.get("/api/kg/thought/{thought_id}/profile")
async def kg_thought_profile(thought_id: str, request: Request):
    """Interpretation: the stored TriBE bridge-profile for a thought.

    Returns the 8-bridge activation levels plus a human-readable summary
    ("high social + memory, low defense"). Populated only when thoughts were
    ingested with TRIBE_PROFILE_ENABLED=1.
    """
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse({"error": "knowledge graph unavailable"}, status_code=503)
    try:
        prof = kg.get_thought_profile(thought_id)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)
    if prof is None:
        return JSONResponse({"error": "thought not found"}, status_code=404)
    if not prof.get("bridge_levels"):
        prof["note"] = "no neural profile stored (TRIBE_PROFILE_ENABLED off at ingest?)"
    return JSONResponse(prof)


@router.get("/api/execution-log/search")
async def execution_log_search(request: Request, q: str = "", diff: str = "",
                               source: str = "", limit: int = 20):
    """RAG over the execution trace (Baustein D.2).

    Query params:
      q      — semantic query (free text); empty → recent/any
      diff   — filter MATCH | MISMATCH | UNVERIFIED (claimed-vs-verified)
      source — filter planner | executor | validator
      limit  — max hits

    Example: /api/execution-log/search?diff=MISMATCH → actions that claimed
    success but the world didn't confirm. Only populated with EXECUTION_LOG_ENABLED=1.
    """
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse({"error": "knowledge graph unavailable"}, status_code=503)
    try:
        from core.execution_log import ExecutionLog, EXECUTION_LOG_ENABLED
        if not EXECUTION_LOG_ENABLED:
            return JSONResponse({
                "enabled": False,
                "note": "set EXECUTION_LOG_ENABLED=1 to record + query the execution trace",
                "results": [],
            })
        log = ExecutionLog(kg)
        hits = log.search(q or "execution step", diff=diff or None,
                          source=source or None, limit=limit)
        return JSONResponse({
            "enabled": True,
            "query": {"q": q, "diff": diff, "source": source, "limit": limit},
            "count": len(hits),
            "results": hits,
        })
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.get("/api/sequence-learner/status")
async def sequence_learner_status(request: Request, intent: str = "", task_type: str = ""):
    """Baustein A — learned agent-sequence stats + optional suggestion.

    Pass ?intent=... or ?task_type=... to see what sequence the learner would
    suggest. Empty → just the learner state. Populated with SEQUENCE_LEARNER_ENABLED=1.
    """
    try:
        from core.sequence_learner import get_learner, SEQUENCE_LEARNER_ENABLED
        learner = get_learner()
        out = {"enabled": SEQUENCE_LEARNER_ENABLED, "state": learner.get_state()}
        if intent or task_type:
            out["suggestion"] = learner.suggest(intent=intent, task_type=task_type)
        return JSONResponse(out)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.get("/api/brain/auto_dispatch_stats")
async def auto_dispatch_stats(request: Request):
    """Stats about Phase F.4 AutoDispatcher (BrainChat -> Minibook)."""
    ad = getattr(request.app.state, "auto_dispatcher", None)
    if ad is None:
        return JSONResponse(
            {"enabled": False, "message": "auto_dispatcher not wired"},
        )
    return JSONResponse({
        "enabled": True,
        "stats": convert_numpy(dict(ad.stats)),
        "timestamp": time.time(),
    })


# ──────────────────────────────────────────────────────────────────────
# Ideas-Space proxy (Phase O.1) — Brain forwards to local Ideas HTTP
# wrapper at port 5102 via state.ideas_client.
# ──────────────────────────────────────────────────────────────────────


def _ideas(request: Request):
    return getattr(request.app.state, "ideas_client", None)


@router.get("/api/ideas/health")
async def ideas_health(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse(
            {"enabled": False, "message": "ideas_client not wired"},
            status_code=503,
        )
    return JSONResponse(ic.health())


@router.get("/api/ideas/list")
async def ideas_list(
    request: Request,
    limit: int = 20,
    query: str = "",
    parent_id: str = "",
):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.list_ideas(
        limit=limit,
        query=query or None,
        parent_id=parent_id if parent_id != "" else None,
    ))


@router.post("/api/ideas/create")
async def ideas_create(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    title = (body.get("title") or "").strip()
    if not title:
        return JSONResponse({"error": "title required"}, status_code=400)
    return JSONResponse(ic.create_idea(
        title=title,
        content=body.get("content") or body.get("description") or "",
        tags=body.get("tags"),
        parent_id=body.get("parent_id"),
        source=body.get("source") or "brain",
    ))


@router.post("/api/ideas/search")
async def ideas_search(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    q = (body.get("query") or "").strip()
    if not q:
        return JSONResponse({"error": "query required"}, status_code=400)
    return JSONResponse(ic.search_ideas(
        query=q,
        limit=int(body.get("limit") or 10),
        min_score=float(body.get("min_score") or 0.3),
    ))


@router.post("/api/ideas/{idea_id}/expand")
async def ideas_expand(idea_id: str, request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    return JSONResponse(ic.expand_idea(
        idea_id=idea_id,
        prompt=body.get("prompt"),
        count=int(body.get("count") or 3),
    ))


# Bubbles (Block 1)


@router.get("/api/bubbles/list")
async def bubbles_list(request: Request, limit: int = 50):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.list_bubbles(limit=limit))


@router.post("/api/bubbles/create")
async def bubbles_create(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    title = (body.get("title") or "").strip()
    if not title:
        return JSONResponse({"error": "title required"}, status_code=400)
    return JSONResponse(ic.create_bubble(
        title=title,
        description=body.get("description") or body.get("content") or "",
        tags=body.get("tags"),
        source=body.get("source") or "brain",
    ))


@router.delete("/api/bubbles/{bubble_id}")
async def bubbles_delete(request: Request, bubble_id: str, force: bool = False):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.delete_bubble(bubble_id, force=force))


@router.post("/api/ideas/{idea_id}/move")
async def ideas_move(idea_id: str, request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    parent_id = body.get("parent_id")
    if parent_id == "":
        parent_id = None
    return JSONResponse(ic.move_idea(idea_id=idea_id, parent_id=parent_id))


# Phase Q.5 — Brain-side proxies for Ideas mini-brain endpoints


@router.get("/api/ideas/kg_stats")
async def ideas_kg_stats(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.kg_stats())


@router.post("/api/ideas/kg_search")
async def ideas_kg_search(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    q = (body.get("query") or "").strip()
    if not q:
        return JSONResponse({"error": "query required"}, status_code=400)
    return JSONResponse(ic.kg_search(
        query=q,
        limit=int(body.get("limit") or 10),
        threshold=float(body.get("threshold") or 0.3),
        node_type=body.get("node_type"),
    ))


@router.get("/api/ideas/state")
async def ideas_state(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.state())


@router.post("/api/ideas/sync_full")
async def ideas_sync_full(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.sync_full())


@router.post("/api/ideas/consolidate")
async def ideas_consolidate(request: Request):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.consolidate_now())


@router.get("/api/ideas/consolidate/suggestions")
async def ideas_consolidate_suggestions(
    request: Request, status: str = "pending", limit: int = 20,
):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.consolidate_suggestions(status=status, limit=limit))


@router.post("/api/ideas/consolidate/suggestions/{suggestion_id}/accept")
async def ideas_consolidate_accept(request: Request, suggestion_id: str):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.consolidate_accept(suggestion_id))


@router.post("/api/ideas/consolidate/suggestions/{suggestion_id}/reject")
async def ideas_consolidate_reject(request: Request, suggestion_id: str):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    return JSONResponse(ic.consolidate_reject(suggestion_id))


@router.post("/api/ideas/{idea_id}/reward")
async def ideas_reward(request: Request, idea_id: str):
    ic = _ideas(request)
    if ic is None:
        return JSONResponse({"error": "ideas_client not wired"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    return JSONResponse(ic.record_reward(
        idea_id=idea_id,
        delta=float(body.get("delta") or 0.0),
        reason=body.get("reason") or "",
    ))


# ─────────────────────────────────────────────────────────────────────
# Phase R — Self-Discourse + Aggregator + Mirofish-KG-Sync
# ─────────────────────────────────────────────────────────────────────


@router.get("/api/discourse/stats")
async def discourse_stats(request: Request):
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"enabled": False, "message": "discourse_engine not running"})
    return JSONResponse(de.stats_dict())


@router.post("/api/discourse/tick_now")
async def discourse_tick_now(request: Request):
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"}, status_code=503)
    return JSONResponse(de.tick_once())


@router.post("/api/discourse/pause")
async def discourse_pause(request: Request):
    """Pause idle + response loops. Intent on-demand still works.
    Useful while Mirofish-Sim is being set up to stop interview-poll spam."""
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"}, status_code=503)
    de.pause()
    return JSONResponse({"ok": True, "paused": True})


@router.post("/api/discourse/resume")
async def discourse_resume(request: Request):
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"}, status_code=503)
    de.resume()
    return JSONResponse({"ok": True, "paused": False})


@router.get("/api/discourse/aggregate_stats")
async def discourse_agg_stats(request: Request):
    agg = getattr(request.app.state, "discourse_aggregator", None)
    if agg is None:
        return JSONResponse({"enabled": False, "message": "aggregator not running"})
    return JSONResponse(agg.stats_dict())


@router.post("/api/discourse/aggregate_now")
async def discourse_aggregate_now(request: Request):
    agg = getattr(request.app.state, "discourse_aggregator", None)
    if agg is None:
        return JSONResponse({"error": "aggregator not running"}, status_code=503)
    return JSONResponse(agg.tick_once())


@router.get("/api/mirofish/sync_stats")
async def mirofish_sync_stats(request: Request):
    mfs = getattr(request.app.state, "mirofish_kg_sync", None)
    if mfs is None:
        return JSONResponse({"enabled": False, "message": "mirofish_kg_sync not running"})
    return JSONResponse(mfs.stats_dict())


@router.post("/api/mirofish/sync_now")
async def mirofish_sync_now(request: Request):
    mfs = getattr(request.app.state, "mirofish_kg_sync", None)
    if mfs is None:
        return JSONResponse({"error": "mirofish_kg_sync not running"}, status_code=503)
    return JSONResponse(mfs.tick_once())


# ─────────────────────────────────────────────────────────────────────
# Phase S.4 — Self-Awareness Watcher
# ─────────────────────────────────────────────────────────────────────


@router.get("/api/self_awareness/manifest_stats")
async def self_awareness_stats(request: Request):
    saw = getattr(request.app.state, "self_awareness_watcher", None)
    if saw is None:
        return JSONResponse({"enabled": False, "message": "watcher not running"})
    return JSONResponse(saw.stats_dict())


@router.post("/api/self_awareness/reseed")
async def self_awareness_reseed(request: Request):
    """Trigger an immediate self-awareness reseed pass.
    Returns {checked, unchanged, updated, added, removed}."""
    saw = getattr(request.app.state, "self_awareness_watcher", None)
    if saw is None:
        return JSONResponse({"error": "watcher not running"}, status_code=503)
    return JSONResponse(saw.tick_once())


# ─────────────────────────────────────────────────────────────────────
# Phase S.5 — Discourse Memory Consolidator (cross-session)
# ─────────────────────────────────────────────────────────────────────


@router.get("/api/discourse/meta_stats")
async def discourse_meta_stats(request: Request):
    dmc = getattr(request.app.state, "discourse_memory_consolidator", None)
    if dmc is None:
        return JSONResponse({"enabled": False, "message": "consolidator not running"})
    return JSONResponse(dmc.stats_dict())


@router.post("/api/discourse/meta_consolidate_now")
async def discourse_meta_consolidate_now(request: Request):
    """Force one cross-session meta-consolidation pass over aggregated-kg.
    Clusters topics, synthesises meta_topics. Returns delta-dict."""
    dmc = getattr(request.app.state, "discourse_memory_consolidator", None)
    if dmc is None:
        return JSONResponse({"error": "consolidator not running"}, status_code=503)
    return JSONResponse(dmc.run_once())


@router.post("/api/self_awareness/recall")
async def self_awareness_recall(request: Request):
    """Recall meta_topics + topics from aggregated-kg matching the query.
    Body: {"query": "...", "days": 7, "limit": 10}"""
    dmc = getattr(request.app.state, "discourse_memory_consolidator", None)
    if dmc is None:
        return JSONResponse({"error": "consolidator not running"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    query = body.get("query") or ""
    days = int(body.get("days") or 7)
    limit = int(body.get("limit") or 10)
    if not query.strip():
        return JSONResponse({"error": "query required"}, status_code=400)
    return JSONResponse(dmc.recall(query, days=days, limit=limit))


@router.get("/api/fungus/stats")
async def fungus_stats(request: Request):
    """Phase S.3 — fungus client status: online flag, doc count, query
    counter, last error."""
    fc = getattr(request.app.state, "fungus_client", None)
    if fc is None:
        return JSONResponse({"online": False, "message": "fungus_client not initialised"})
    return JSONResponse(fc.stats_dict())


@router.get("/api/capabilities/stats")
async def capability_stats(request: Request):
    """Phase 1 — capability router state: registry size, match counters,
    list of loaded capabilities."""
    cr = getattr(request.app.state, "capability_router", None)
    if cr is None:
        return JSONResponse({
            "loaded": False,
            "message": "capability_router not initialised",
        })
    return JSONResponse({"loaded": True, **cr.stats_dict()})


@router.get("/api/capabilities/list")
async def capability_list(request: Request):
    """Phase 1 — show all loaded capabilities with their primary/supporting
    agents and execution targets, useful for debugging registry-rot."""
    cr = getattr(request.app.state, "capability_router", None)
    if cr is None:
        return JSONResponse({"loaded": False, "capabilities": []})
    return JSONResponse({"loaded": True, "capabilities": cr.list_capabilities()})


@router.get("/api/capabilities/by_name/{name}")
async def capability_detail(name: str, request: Request):
    """Show full detail of a single capability — its patterns, anchor
    phrases, agents, execution target, embedding status. Useful when
    debugging why an intent did or didn't match a specific capability.

    Path uses /by_name/ prefix to avoid clashing with literal sub-routes
    like /api/capabilities/targets, /api/capabilities/validator/stats etc."""
    cr = getattr(request.app.state, "capability_router", None)
    if cr is None:
        return JSONResponse({"error": "capability_router not loaded"}, status_code=503)
    detail = cr.get_capability(name)
    if detail is None:
        return JSONResponse({"error": f"capability '{name}' not found"}, status_code=404)
    return JSONResponse(detail)


@router.post("/api/capabilities/reload")
async def capability_reload(request: Request):
    """Phase 1+2 — re-read data/capabilities.yaml without restarting Brain.
    Re-builds embeddings if the embedder is wired. Useful for iterating on
    YAML pattern + description tuning.

    Phase 11.M — `cr.reload()` is sync and CPU-heavy (re-embeds 100+ anchors
    via SentenceTransformer). Running it inline blocks the asyncio event
    loop for 30-60s on first reload, freezing every other endpoint. Move it
    to a worker thread so other requests keep flowing.
    """
    import asyncio
    cr = getattr(request.app.state, "capability_router", None)
    if cr is None:
        return JSONResponse({"error": "capability_router not loaded"}, status_code=503)
    try:
        await asyncio.to_thread(cr.reload)
        return JSONResponse({"ok": True, **cr.stats_dict()})
    except Exception as e:
        return JSONResponse(
            {"ok": False, "error": f"{type(e).__name__}: {e}"},
            status_code=500,
        )


@router.post("/api/capabilities/test")
async def capability_test(request: Request):
    """Phase 1 — test the router without running discourse. Body:
    {"intent": "..."}. Returns the match (or no-match) so the YAML
    registry can be debugged without hitting the full discourse stack.

    Phase 5 — also feeds the curator's telemetry so /test can be used
    to populate the cluster pool from outside DiscourseEngine."""
    cr = getattr(request.app.state, "capability_router", None)
    if cr is None:
        return JSONResponse({"error": "capability_router not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    intent = (body.get("intent") or body.get("message") or "").strip()
    if not intent:
        return JSONResponse({"error": "intent required"}, status_code=400)
    m = cr.route(intent)

    # Phase 5 — record into curator if wired
    cur = getattr(request.app.state, "capability_curator", None)
    if cur is not None:
        try:
            cur.record_intent(
                intent,
                matched=bool(m),
                capability=m.capability if m else None,
                match_method=m.match_method if m else None,
            )
        except Exception:
            pass

    if m is None:
        return JSONResponse({"matched": False, "intent": intent})
    return JSONResponse({
        "matched": True,
        "intent": intent,
        "capability": m.capability,
        "description": m.description,
        "primary": m.primary_names,
        "supporting": m.supporting_names,
        "matched_pattern": m.matched_pattern,
        "match_method": m.match_method,
        "is_direct": m.is_direct,
        "execution_target": m.execution_target,
        "validator": m.validator,
    })


# ── Phase 3: Validator endpoints ──────────────────────────────────────


@router.get("/api/capabilities/validator/stats")
async def validator_stats(request: Request):
    """Phase 3 — validator activity counters."""
    cv = getattr(request.app.state, "capability_validator", None)
    if cv is None:
        return JSONResponse({"loaded": False, "message": "validator not initialised"})
    return JSONResponse({"loaded": True, **cv.stats_dict()})


@router.post("/api/capabilities/validator/test")
async def validator_test(request: Request):
    """Phase 3 — test a validator config without running discourse. Body:
        {"validator": {...}, "intent": "...", "raw_result": {...}}
    Returns the verdict envelope (valid, reason, kind, on_fail, elapsed_s).
    """
    cv = getattr(request.app.state, "capability_validator", None)
    if cv is None:
        return JSONResponse({"error": "validator not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    validator_cfg = body.get("validator") or {}
    intent = (body.get("intent") or "").strip()
    raw_result = body.get("raw_result")
    arg = body.get("arg")
    if not validator_cfg or not isinstance(validator_cfg, dict):
        return JSONResponse({"error": "validator config required"}, status_code=400)
    verdict = cv.validate(validator_cfg, intent=intent, arg=arg, raw_result=raw_result)
    return JSONResponse(verdict)


# ── Phase 4: Execution-targets introspection ──────────────────────────


@router.get("/api/capabilities/targets")
async def capability_targets(request: Request):
    """Phase 4 — show which target kinds are supported and per-target stats
    for any executors that have been instantiated so far."""
    try:
        from core.capability_targets import supported_kinds
        kinds = supported_kinds()
    except Exception as e:
        return JSONResponse({"error": f"targets module unavailable: {e}"}, status_code=503)
    de = getattr(request.app.state, "discourse_engine", None)
    cache = {}
    if de is not None and hasattr(de, "_executor_cache") and de._executor_cache:
        for tgt, exe in de._executor_cache.items():
            if exe is None:
                cache[tgt] = {"target": tgt, "resolvable": False}
                continue
            try:
                cache[tgt] = exe.stats_dict()
            except Exception as e:
                cache[tgt] = {"target": tgt, "error": str(e)}
    return JSONResponse({
        "supported_kinds": kinds,
        "instantiated_executors": list(cache.values()),
    })


# ── Phase 5: Curator endpoints ────────────────────────────────────────


@router.get("/api/capabilities/curator/stats")
async def curator_stats(request: Request):
    """Phase 5 — curator counters: intents logged, no-matches, suggestions
    generated/accepted/rejected, last cluster run."""
    cur = getattr(request.app.state, "capability_curator", None)
    if cur is None:
        return JSONResponse({"loaded": False, "message": "curator not initialised"})
    return JSONResponse({"loaded": True, **cur.stats_dict()})


@router.post("/api/capabilities/curator/suggest")
async def curator_suggest(request: Request):
    """Phase 5 — run a clustering pass over no-match intents and produce
    suggestions for new capabilities. Body (optional):
        {"max_suggestions": 5, "min_age_s": 0}
    Returns the list of pending suggestion dicts."""
    cur = getattr(request.app.state, "capability_curator", None)
    if cur is None:
        return JSONResponse({"error": "curator not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    max_n = int(body.get("max_suggestions") or 5)
    min_age = float(body.get("min_age_s") or 0)
    suggestions = cur.suggest(max_suggestions=max_n, min_age_s=min_age)
    return JSONResponse({
        "ok": True,
        "generated": len(suggestions),
        "suggestions": suggestions,
    })


@router.get("/api/capabilities/curator/suggestions")
async def curator_list_suggestions(request: Request):
    """Phase 5 — list all suggestions, optional ?status=pending|accepted|rejected."""
    cur = getattr(request.app.state, "capability_curator", None)
    if cur is None:
        return JSONResponse({"error": "curator not loaded"}, status_code=503)
    status_filter = request.query_params.get("status")
    items = cur.list_suggestions(status=status_filter)
    return JSONResponse({"count": len(items), "suggestions": items})


@router.post("/api/capabilities/curator/suggestions/{suggestion_id}/accept")
async def curator_accept(suggestion_id: str, request: Request):
    """Phase 5 — accept a suggestion. Optional body lets the user pin
    agent names or an execution target before the YAML write:
        {"agents_primary": [...], "agents_supporting": [...],
         "execution_target": "...", "reason": "..."}"""
    cur = getattr(request.app.state, "capability_curator", None)
    if cur is None:
        return JSONResponse({"error": "curator not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    out = cur.accept(
        suggestion_id,
        agents_primary=body.get("agents_primary"),
        agents_supporting=body.get("agents_supporting"),
        execution_target=body.get("execution_target"),
        reason=body.get("reason"),
    )
    code = 200 if out.get("ok") else 400
    return JSONResponse(out, status_code=code)


@router.post("/api/capabilities/curator/suggestions/{suggestion_id}/reject")
async def curator_reject(suggestion_id: str, request: Request):
    """Phase 5 — reject a suggestion."""
    cur = getattr(request.app.state, "capability_curator", None)
    if cur is None:
        return JSONResponse({"error": "curator not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    out = cur.reject(suggestion_id, reason=body.get("reason"))
    code = 200 if out.get("ok") else 400
    return JSONResponse(out, status_code=code)


# ── Phase 6: Multi-hop endpoints ──────────────────────────────────────


@router.get("/api/multihop/provider_scores")
async def multihop_provider_scores(request: Request):
    """Phase 7.3 — per-(capability, target-kind) success rate. Adaptive
    Routing reads this to break ties when multiple providers can serve a
    capability."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"loaded": False, "scores": {}})
    return JSONResponse({"loaded": True, "scores": pe.get_provider_scores()})


@router.post("/api/multihop/plan/{plan_id}/reward")
async def multihop_plan_reward(plan_id: str, request: Request):
    """Phase 7.1 — manually attach a reward delta to a plan. Body: {delta, reason}."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"error": "plan_executor not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    delta = float(body.get("delta") or 0.0)
    reason = body.get("reason") or "manual"
    out = pe.record_plan_reward(plan_id, delta, reason=reason)
    code = 200 if out.get("ok") else 404
    return JSONResponse(out, status_code=code)


# ── Phase 8: Cluster Engine endpoints ───────────────────────────────


@router.get("/api/clusters/activations")
async def clusters_activations(request: Request):
    """Phase 8.1 — per-cluster activation snapshot used by Galaxy UI
    + Self-Steerer."""
    ce = getattr(request.app.state, "cluster_engine", None)
    if ce is None:
        return JSONResponse({"loaded": False, "clusters": []})
    return JSONResponse({"loaded": True, "clusters": ce.get_activations()})


@router.get("/api/clusters/co_activations")
async def clusters_co_activations(request: Request):
    """Phase 8.1 — pairs of clusters that co-activated recently
    (renders as edges in the galaxy UI)."""
    ce = getattr(request.app.state, "cluster_engine", None)
    if ce is None:
        return JSONResponse({"loaded": False, "pairs": []})
    return JSONResponse({"loaded": True, "pairs": ce.get_co_activations()})


@router.get("/api/clusters/stats")
async def clusters_stats(request: Request):
    """Phase 8.1 — engine ticker / decay / running state."""
    ce = getattr(request.app.state, "cluster_engine", None)
    if ce is None:
        return JSONResponse({"loaded": False})
    return JSONResponse({"loaded": True, **ce.stats_dict()})


@router.post("/api/clusters/bump")
async def clusters_bump(request: Request):
    """Phase 8.1 — manually nudge a cluster's activation. Used by smoke
    tests and a 'fire cluster' button in the UI. Body: {cluster_id, delta}."""
    ce = getattr(request.app.state, "cluster_engine", None)
    if ce is None:
        return JSONResponse({"error": "cluster_engine not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    cid = (body.get("cluster_id") or "").strip()
    delta = float(body.get("delta") or 0.0)
    if not cid:
        return JSONResponse({"error": "cluster_id required"}, status_code=400)
    out = ce.bump(cid, delta)
    code = 200 if out.get("ok") else 404
    return JSONResponse(out, status_code=code)


@router.post("/api/clusters/tick_now")
async def clusters_tick_now(request: Request):
    """Phase 8.1 — manual one-shot tick (for tests; usually runs auto every 60s)."""
    ce = getattr(request.app.state, "cluster_engine", None)
    if ce is None:
        return JSONResponse({"error": "cluster_engine not loaded"}, status_code=503)
    import asyncio as _asyncio
    loop = _asyncio.get_running_loop()
    out = await loop.run_in_executor(None, ce.tick_once)
    return JSONResponse(out)


# ── Phase 8.B: Decision-Graph endpoints (Neo4j) ──────────────────────


@router.get("/api/decision_graph/query")
async def decision_graph_query(request: Request, limit: int = 200, min_activation: float = 0.05):
    """Phase 8.B — Cytoscape-formatted subgraph for the decision-theatre UI.

    Filters by min_activation (only show clusters above this threshold).
    Returns recent plans/hops/dispatches from the last 24h."""
    dg = getattr(request.app.state, "decision_graph", None)
    if dg is None:
        return JSONResponse({"connected": False, "nodes": [], "edges": []})
    import asyncio as _asyncio
    loop = _asyncio.get_running_loop()
    out = await loop.run_in_executor(
        None, dg.query_subgraph, int(limit), float(min_activation),
    )
    return JSONResponse(out)


@router.get("/api/decision_graph/stats")
async def decision_graph_stats(request: Request):
    """Phase 8.B — counts per label + connection state."""
    dg = getattr(request.app.state, "decision_graph", None)
    if dg is None:
        return JSONResponse({"connected": False})
    return JSONResponse(dg.stats())


# ── Phase 9.0.4: Approval-Gate endpoints ──────────────────────────


@router.post("/api/approvals/{tool_call_id}")
async def approve_tool_call(tool_call_id: str, request: Request):
    """Phase 9.0.4 — user clicks Approve/Deny in the UI modal. Body:
        {"decision": "approve" | "deny"}
    Stores decision in-memory + writes approval_status to Neo4j ToolCall."""
    ag = getattr(request.app.state, "approval_gate", None)
    if ag is None:
        return JSONResponse({"error": "approval_gate not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        body = {}
    decision = (body.get("decision") or "").strip().lower()
    if decision not in ("approve", "deny"):
        return JSONResponse({"error": "decision must be 'approve' or 'deny'"}, status_code=400)
    out = ag.submit(tool_call_id, decision)
    return JSONResponse(out)


@router.get("/api/approvals/stats")
async def approvals_stats(request: Request):
    """Phase 9.0.4 — counts of approved/denied/requested decisions."""
    ag = getattr(request.app.state, "approval_gate", None)
    if ag is None:
        return JSONResponse({"loaded": False})
    return JSONResponse({"loaded": True, **ag.stats_dict()})


@router.post("/api/decision_graph/prune")
async def decision_graph_prune(request: Request, older_than_s: float = 604800):
    """Phase 8.B — drop ancient nodes. Default TTL 7d."""
    dg = getattr(request.app.state, "decision_graph", None)
    if dg is None:
        return JSONResponse({"error": "decision_graph not loaded"}, status_code=503)
    import asyncio as _asyncio
    loop = _asyncio.get_running_loop()
    out = await loop.run_in_executor(None, dg.prune, float(older_than_s))
    return JSONResponse(out)


# ── Phase 8.3: Self-Steerer endpoints ───────────────────────────────


@router.get("/api/self_steer/stats")
async def self_steer_stats(request: Request):
    """Phase 8.3 — autonomous-dispatch counters + last decision."""
    ss = getattr(request.app.state, "self_steerer", None)
    if ss is None:
        return JSONResponse({"loaded": False})
    return JSONResponse({"loaded": True, **ss.stats_dict()})


@router.post("/api/self_steer/tick_now")
async def self_steer_tick_now(request: Request):
    """Phase 8.3 — manual scan-and-dispatch (skip 30s wait)."""
    ss = getattr(request.app.state, "self_steerer", None)
    if ss is None:
        return JSONResponse({"error": "self_steerer not loaded"}, status_code=503)
    import asyncio as _asyncio
    loop = _asyncio.get_running_loop()
    out = await loop.run_in_executor(None, ss.tick_once)
    return JSONResponse(out)


@router.post("/api/self_steer/reload_mappings")
async def self_steer_reload(request: Request):
    """Phase 8.3 — reload cluster_capabilities.yaml live (no Brain restart)."""
    ss = getattr(request.app.state, "self_steerer", None)
    if ss is None:
        return JSONResponse({"error": "self_steerer not loaded"}, status_code=503)
    n = ss.reload_mappings()
    return JSONResponse({"ok": True, "mappings_loaded": n})


@router.get("/api/multihop/busy")
async def multihop_busy(request: Request):
    """Phase 6.14.1 — is a plan currently running? UI uses this to
    enable/disable the Plan+Execute button. Cheap check."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"loaded": False, "busy": False})
    return JSONResponse({"loaded": True, **pe.busy_status()})


@router.get("/api/multihop/stats")
async def multihop_stats(request: Request):
    """Phase 6 — combined counters from advisor, planner, executor, synthesizer."""
    state = request.app.state
    out: Dict[str, Any] = {"loaded": False}
    parts: Dict[str, Any] = {}
    adv = getattr(state, "multihop_advisor", None)
    if adv is not None:
        parts["advisor"] = adv.stats_dict()
    pl = getattr(state, "multihop_planner", None)
    if pl is not None:
        parts["planner"] = pl.stats_dict()
    pe = getattr(state, "plan_executor", None)
    if pe is not None:
        parts["executor"] = pe.stats_dict()
    syn = getattr(state, "final_synthesizer", None)
    if syn is not None:
        parts["synthesizer"] = syn.stats_dict()
    if parts:
        out = {"loaded": True, **parts}
    return JSONResponse(out)


@router.post("/api/multihop/plan")
async def multihop_plan_only(request: Request):
    """Phase 6 — produce a plan without executing it. Body: {intent}."""
    pl = getattr(request.app.state, "multihop_planner", None)
    if pl is None:
        return JSONResponse({"error": "planner not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    intent = (body.get("intent") or body.get("message") or "").strip()
    if not intent:
        return JSONResponse({"error": "intent required"}, status_code=400)

    # Phase 11.T.4 — prefer the async path so we don't burn a threadpool
    # worker on the LLM round-trip. Falls back to threadpool wrapper only
    # if planner instance predates the aplan() addition.
    aplan = getattr(pl, "aplan", None)
    if aplan is not None:
        plan = await aplan(intent)
    else:
        import asyncio as _asyncio
        loop = _asyncio.get_running_loop()
        plan = await loop.run_in_executor(None, pl.plan, intent)
    if plan is None:
        return JSONResponse({"ok": False, "error": pl.stats_dict().get("last_error")})
    return JSONResponse({"ok": True, "plan": plan.to_dict()})


# ─── Phase 11.Q2 — Capability-Router shortcut helper ───────────────────

# Multi-action signal: count distinct command-verbs. Two+ verbs = LLM path
# (multi-hop plan). One verb = single-action, eligible for shortcut.
_SHORTCUT_VERB_RE = re.compile(
    # Note: "format" is BOTH a verb and a common noun ("table-format",
    # "the format of"). Excluded from multi-action detection — better
    # to under-count and shortcut more often than to over-count and
    # send everything to LLM.
    r"\b(create|delete|remove|add|update|find|search|list|show|"
    r"connect|disconnect|trenne|verbinde|wandle|mach(?:e)?|"
    r"erstelle|leg(?:e)?|loesche|lösche|"
    r"benenne|umbenenne|geh|enter|verlasse|exit|"
    r"explain|erklaere|erkläre|expand|erweitere|"
    r"score|rate|bewerte|evaluate|analyze|analysiere|untersuche|"
    r"rebuild|regenerate|reindex)\b",
    re.IGNORECASE,
)


def _looks_like_multi_action(intent_lower: str) -> bool:
    """Heuristic: ≥2 distinct command-verbs → multi-action (use LLM).
    Note: 'und' between items (e.g. 'trenne A und B') is single-action,
    so we count verbs, not 'und' occurrences."""
    verbs = set(_SHORTCUT_VERB_RE.findall(intent_lower))
    if len(verbs) >= 2:
        return True
    if " then " in intent_lower or " dann " in intent_lower:
        return True
    if intent_lower.count(",") >= 2:  # "A, B, C" style batch
        return True
    if ";" in intent_lower:
        return True
    return False


def _extract_arg_template(intent: str, capability: str) -> str:
    """Extract the most plausible argument from the intent string.
    Quoted spans win; otherwise prefer last capitalised/identifier-like
    token. Returns empty string when we can't find a candidate."""
    # 1. Quoted span ("foo bar" or 'foo bar')
    qm = re.search(r"['\"]([^'\"]+)['\"]", intent)
    if qm:
        return qm.group(1).strip()
    # 2. Last capitalised or underscore-containing token
    tokens = re.findall(r"\b[\w_-]+\b", intent)
    if not tokens:
        return ""
    for tok in reversed(tokens):
        if tok and (tok[0].isupper() or "_" in tok):
            # Skip stop-words even if capitalised
            if tok.lower() in {"a", "an", "the", "und", "and", "die", "der", "das"}:
                continue
            return tok
    # 3. Fall back to last token
    return tokens[-1]


def _try_capability_shortcut(state, intent: str):
    """Phase 11.Q2 — try to short-circuit the LLM planner for trivial
    single-action intents.

    Returns a `Plan` object on success, `None` to fall through to the
    LLM planner. Never raises (wraps everything in try/except).

    Strategy:
      - Multi-action intents (verb count ≥ 2) → None (use LLM)
      - Capability-Router regex match + is_direct → 1-hop plan
      - Anything else → None (LLM gets a chance to disambiguate)
    """
    cr = getattr(state, "capability_router", None)
    if cr is None:
        return None

    intent_lower = intent.lower()
    if _looks_like_multi_action(intent_lower):
        logger.debug(f"[shortcut] multi-action intent, deferring to LLM: {intent!r}")
        return None

    # Route via capability-router
    try:
        m = cr.route(intent)
    except Exception as e:
        logger.warning(f"[shortcut] cr.route crashed: {e}")
        return None
    if m is None:
        return None

    # Only accept regex matches (deterministic). Semantic-pull is too
    # noisy for the shortcut — let the LLM disambiguate those.
    method = getattr(m, "match_method", "")
    if method != "regex":
        logger.debug(f"[shortcut] non-regex match ({method}), deferring to LLM")
        return None

    # Phase 11.U.C — multi-arg caps the shortcut can't extract correctly.
    # Shortcut grabs ONE token from the intent; these need TWO+ (source+target,
    # bubble+name, etc.) and must go through the LLM planner.
    _MULTI_ARG_CAPS = {
        "idea_connect", "idea_disconnect", "idea_connect_multi",
        "bubble_update",  # needs old_name + new_name
        "idea_move",      # needs idea + target_bubble
        "idea_add",       # Phase 11.U.D — needs (title, bubble_name)
        "idea_create",    # same
    }
    if m.capability in _MULTI_ARG_CAPS:
        logger.debug(f"[shortcut] {m.capability!r} is multi-arg, deferring to LLM")
        return None

    # Phase 11.U.K — state-dependent caps. These read voice-process-local
    # `_current_bubble_db_id`, so the MCP-direct dispatch path won't see the
    # bubble context. We must let the LLM planner produce a `bubble_enter`-
    # first multi-hop plan, OR if the intent already names a bubble, the
    # planner emits a 2-hop plan (bubble_enter, then the operation).
    # Single-hop shortcut → empty bubble state → "Please enter a Space first."
    _STATE_DEPENDENT_CAPS = {
        # idea-formatters all call _get_current_bubble_id() internally
        "idea_format_table", "idea_format_note", "idea_format_action_list",
        "idea_format_pros_cons", "idea_format_hierarchy", "idea_format_specs",
        "idea_format_kanban", "idea_format_mindmap", "idea_format_swot",
        "idea_format_user_story", "idea_format_flowchart",
        "idea_convert_format", "idea_format_revert",
        "idea_format_get", "idea_format_list",
        # content-tools that work on "the idea in the current bubble"
        "idea_explain", "idea_classify", "idea_expand", "idea_update",
        "idea_find", "idea_count", "idea_list", "idea_delete",
        "idea_link_to_root",
        # analysis-tools that scan current bubble
        "idea_auto_link", "idea_analyze_links",
        "bubble_generate_embeddings",
        # bubble lifecycle inside-context
        "bubble_exit",
    }
    # State-dependent caps need a bubble context. If the intent already names
    # a bubble OR a specific idea, defer to the LLM planner — it'll produce
    # bubble_enter (or idea_find→bubble) as the first hop. Otherwise we'd
    # dispatch into an empty voice-process bubble state and get "Please
    # enter a Space first."
    intent_lower = intent.lower()
    # Bubble mentions: cover dative + accusative + English
    mentions_bubble = any(
        kw in intent_lower for kw in (
            "bubble ", "space ", "raum ", "blase ",
            "geh in", "geh zur", "betrete", "go to ", "enter ",
        )
    )
    # Idea mentions: "die idee X", "formatiere X", "expand idea X", etc.
    # An all-caps or PascalCase token (≥ 2 chars + at least 1 underscore OR
    # uppercase letter) is a strong signal that the intent names a specific idea.
    import re as _re
    has_named_idea = bool(
        _re.search(r"\b[A-Z][A-Za-z0-9_]*[_A-Z][A-Za-z0-9_]*\b", intent)
        or any(kw in intent_lower for kw in ("die idee ", "the idea ", "idea "))
    )
    if m.capability in _STATE_DEPENDENT_CAPS and (mentions_bubble or has_named_idea):
        logger.info(
            f"[shortcut] {m.capability!r} is state-dependent and intent mentions "
            f"bubble={mentions_bubble} named_idea={has_named_idea} — deferring to LLM"
        )
        return None

    # Must have a registered execution_target so an executor can run the hop
    # without the LLM planner. Phase 11.Q2 originally restricted this to
    # `direct:` (in-process Python) caps; but `openfang:`, `coding-engine:`,
    # `http:`, `n8n:`, `mcp:` executors resolve their own targets just as well
    # (e.g. OpenFangExecutor looks up the agent via /api/agents). Restricting
    # to is_direct sent every openfang:/openclaw cap to the LLM planner, which
    # then hallucinated bubble-creation plans out of a recall saturated with
    # old bubble_create plans. Accept any registered execution-target kind.
    if not getattr(m, "has_execution_target", False):
        return None

    # Look up arg_kwarg from the registry (arg_template gets extracted
    # from intent text). Use empty defaults if anything fails.
    arg_kwarg = None
    try:
        cap_detail = cr.get_capability(m.capability)
        if cap_detail:
            arg_kwarg = cap_detail.get("arg_kwarg")
    except Exception as e:
        logger.debug(f"[shortcut] get_capability failed: {e}")

    arg_template = ""
    if arg_kwarg:
        try:
            arg_template = _extract_arg_template(intent, m.capability)
        except Exception as e:
            logger.debug(f"[shortcut] arg extraction failed: {e}")
            arg_template = ""

    # Build a clean 1-hop plan
    from core.plan_schema import Plan as _Plan, HopSpec as _HopSpec
    try:
        plan = _Plan(
            plan_id=f"shortcut_{int(time.time() * 1000)}",
            intent=intent,
            rationale=f"single-cap shortcut: {m.capability} (regex)",
            hops=[_HopSpec(
                step_id="s1",
                description=f"{m.capability} via shortcut",
                capability=m.capability,
                arg_kwarg=arg_kwarg,
                arg_template=arg_template,
            )],
            final_synthesis_prompt="",
            estimated_cost_usd=0.0,
        )
    except Exception as e:
        logger.warning(f"[shortcut] plan build failed: {e}")
        return None

    logger.info(
        f"[shortcut] {m.capability!r} arg_kwarg={arg_kwarg!r} "
        f"arg={arg_template!r} (regex)"
    )
    return plan


@router.post("/api/multihop/execute")
async def multihop_execute(request: Request):
    """Phase 6 — full intent → plan → execute → synth pipeline. Body:
        {intent}                  produce plan + execute
        {plan: {...}}             execute a hand-built plan (skip planner)
    Returns the executed plan summary plus optional final synthesis."""
    import os  # modulweit nicht importiert in dieser Datei — für SoM-Merge-Routing
    state = request.app.state
    pe = getattr(state, "plan_executor", None)
    pl = getattr(state, "multihop_planner", None)
    syn = getattr(state, "final_synthesizer", None)
    if pe is None:
        return JSONResponse({"error": "plan_executor not loaded"}, status_code=503)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)

    plan_dict = body.get("plan")
    intent = (body.get("intent") or body.get("message") or "").strip()
    skip_shortcut = bool(body.get("force_planner"))

    # E2E-Trace (2026-06-09): EINE durchgaengige Correlation-ID am Eingang. Faedelt
    # durch ALLE Zweige (auch SoM/som-team/no-plan, die kein plan_id haben) - sie
    # steht in jeder Response + (wo ein Plan existiert) in plan.trace_id, sodass
    # GET /api/trace/{trace_id} die ganze Kette eingabe->...->ausgabe zeigt.
    import uuid as _uuid
    trace_id = "tr_" + _uuid.uuid4().hex[:12]
    _routed_via = None   # wird je Zweig gesetzt (groq/som/som-team/no-plan/easy/meta)

    from core.plan_schema import Plan as _Plan
    plan = None
    if plan_dict:
        try:
            plan = _Plan.from_dict(plan_dict)
        except Exception as e:
            return JSONResponse({"error": f"invalid plan: {e}", "trace_id": trace_id}, status_code=400)
    elif intent:
        # Phase 11.Q2 — Capability-Router shortcut. If the router has a
        # high-confidence regex match for a single-action intent, skip
        # the LLM planner and build a 1-hop plan directly. Wrapped in
        # try/except so any bug here falls through to the LLM path
        # instead of 500'ing the whole request.
        if not skip_shortcut:
            try:
                plan = _try_capability_shortcut(state, intent)
            except Exception as e:
                logger.exception(f"[multihop] shortcut crashed for intent={intent!r}: {e}")
                plan = None  # fall through to LLM

        # Phase A (2026-06-08): SEMANTISCHES schwierigkeits-basiertes Routing.
        # Ersetzt das fragile Verb-Zählen (_looks_like_multi_action) durch einen
        # Qwen-Cosine-Klassifikator (difficulty_router): easy→Chat, medium→
        # Shortcut/Groq, hard→SoM, insane→AutoGen(Phase B; bis dahin SoM). Löst
        # den Fall "erstelle Excel" (1 Verb, aber komplex → hard → SoM) der am
        # Verb-Zählen vorbeirutschte. Kill-Switch DIFFICULTY_ROUTING=0 → altes
        # Multi-Action-Verhalten. Klassifikation ist robust gekapselt (Fehler →
        # Heuristik), darf den Handler nie 500'en.
        som_route = False
        if plan is None:
            use_som = os.environ.get("SOM_AS_PLANNER", "1") not in ("0", "false", "False")
            diff_on = os.environ.get("DIFFICULTY_ROUTING", "1") not in ("0", "false", "False")
            level = None
            if diff_on:
                try:
                    from core.difficulty_router import get_router, handler_for
                    # classify() macht einen Qwen-Embedding-Forward-Pass (CPU-schwer)
                    # + beim ersten Call den ~148s-Modell-Cold-Load. SYNC auf dem
                    # async Event-Loop wuerde das ALLE Requests blockieren (jeder
                    # classify pegte einen Core + starvte HTTP — root-caused 2026-06-08).
                    # In den Threadpool offloaden, wie die SoM/som-team-Dispatches.
                    import asyncio as _asyncio
                    _loop = _asyncio.get_running_loop()
                    cls = await _loop.run_in_executor(None, get_router().classify, intent)
                    level = cls.get("level")
                    handler = handler_for(level)
                    logger.info(f"[multihop] difficulty={level} handler={handler} "
                                f"({cls.get('method')}, {cls.get('reason')}) intent={intent[:60]!r}")
                except Exception as e:  # noqa: BLE001 — Klassifikation darf nie 500'en
                    logger.warning(f"[multihop] difficulty classify failed ({e}), Verb-Heuristik")
                    level = None

            if level == "meta":
                # System-/Meta-Nachricht (Konversations-Summary, zurückgespielter
                # Transcript) — KEIN planbarer Intent. NIE an SoM/Groq geben (war
                # Root-Cause des SoM-Run-Storms 2026-06-08). Höflich abweisen.
                logger.info(f"[multihop] meta-Nachricht abgewiesen (kein Plan): {intent[:60]!r}")
                _trace_lite(pe, trace_id, intent, "meta-reject", "")
                return JSONResponse({
                    "ok": True, "difficulty": "meta", "executed": {}, "skipped": True,
                    "final_text": "", "trace_id": trace_id,
                })

            if level == "easy":
                # einfache Frage/Smalltalk → direkte Chat-Antwort, KEIN Planer.
                # brain_chat.send(msg) -> BrainChatResponse (.to_dict()), sync →
                # im Threadpool, damit der Event-Loop frei bleibt.
                bc = getattr(state, "brain_chat", None)
                reply = None
                if bc is not None:
                    try:
                        import asyncio as _asyncio
                        loop = _asyncio.get_running_loop()
                        resp = await loop.run_in_executor(None, bc.send, intent)
                        d = resp.to_dict() if hasattr(resp, "to_dict") else (resp or {})
                        if isinstance(d, dict):
                            reply = d.get("text") or d.get("reply") or d.get("response") or d.get("message")
                        elif isinstance(d, str):
                            reply = d
                    except Exception as e:  # noqa: BLE001
                        logger.warning(f"[multihop] easy-chat failed ({e})")
                _final = (reply.strip() if isinstance(reply, str) and reply.strip() else "Alles klar.")
                _trace_lite(pe, trace_id, intent, "easy-chat", _final)
                return JSONResponse({
                    "ok": True, "difficulty": "easy", "executed": {},
                    "final_text": _final, "trace_id": trace_id,
                })

            if level == "insane" and os.environ.get("INSANE_AUTOGEN", "0") in ("1", "true", "True"):
                # Phase B: vage/explorative Intents → dynamisches AutoGen-Capability-
                # Team (SelectorGroupChat). async wie SoM (Ergebnis per Telegram).
                # Default AUS (INSANE_AUTOGEN=0) → insane fällt unten auf SoM zurück
                # (sicher, solange das Team-Geschütz nicht breit verifiziert ist).
                try:
                    from core.capability_targets import build_executor
                    ex = build_executor("openfang:som-team")
                    # call_with_arg ist SYNC (requests.post an OpenFang). Auf dem
                    # async Event-Loop wuerde das ALLE Requests blockieren
                    # (Handler-Starvation, root-caused 2026-06-08) → in den
                    # Threadpool offloaden, wie der easy/SoM-Pfad.
                    import asyncio as _asyncio
                    _loop = _asyncio.get_running_loop()
                    # _trace_id mitgeben (Phase 2b): der som-team-Worker pusht seine
                    # Stage-Events unter derselben trace_id zurueck an den Trace.
                    res = await _loop.run_in_executor(
                        None, lambda: ex.call_with_arg(intent, extra_params={"_intent": intent, "_trace_id": trace_id}))
                    reply = ""
                    if isinstance(res, dict):
                        reply = res.get("response") or res.get("final_text") or ""
                    _final = (reply.strip() if isinstance(reply, str) and reply.strip()
                              else "An das Multi-Agent-Team übergeben — das Ergebnis kommt per Telegram.")
                    _trace_lite(pe, trace_id, intent, "som-team", _final)
                    return JSONResponse({
                        "ok": True, "difficulty": "insane", "autogen": True, "executed": {},
                        "final_text": _final, "trace_id": trace_id,
                    })
                except Exception as e:  # noqa: BLE001 — Team-Dispatch-Fehler → SoM-Fallback
                    logger.warning(f"[multihop] som-team dispatch failed ({e}), Fallback SoM")

            if use_som and (level in ("hard", "insane") or
                            (level is None and _looks_like_multi_action(intent.lower())) or
                            pl is None):
                # hard → SoM; insane → SoM-Fallback (wenn INSANE_AUTOGEN aus/fehlgeschlagen).
                # level None (Klassifikation aus/fehlgeschlagen) → alte Verb-Heuristik.
                som_route = True  # -> SoM-Dispatch unten, KEIN Groq-Versuch
            else:
                # medium (1 klare Aktion) → Shortcut hat schon gegriffen oder
                # Groq-Multihop versuchen (wie bisher).
                if pl is None:
                    return JSONResponse({"error": "planner not loaded — pass a 'plan' instead"}, status_code=503)
                # Baustein A — inject the learned agent-sequence as a planner prior
                # (no-op unless SEQUENCE_LEARNER_ENABLED). The LLM stays in control;
                # this only hints the proven decomposition.
                _plan_ctx = None
                try:
                    from core.sequence_learner import suggest_sequence
                    _sug = suggest_sequence(intent=intent)
                    if _sug and _sug.get("sequence"):
                        _seq = " → ".join(_sug["sequence"])
                        _plan_ctx = {"hint": (
                            f"A similar intent succeeded {_sug['ok']}x with this agent "
                            f"sequence: {_seq}. Prefer it unless the intent clearly differs."
                        )}
                except Exception:
                    _plan_ctx = None
                # Phase 11.T.4 — async planning path; falls back to threadpool
                aplan_fn = getattr(pl, "aplan", None)
                if aplan_fn is not None:
                    plan = await (aplan_fn(intent, context=_plan_ctx) if _plan_ctx else aplan_fn(intent))
                else:
                    import asyncio as _asyncio
                    loop = _asyncio.get_running_loop()
                    if _plan_ctx:
                        plan = await loop.run_in_executor(None, lambda: pl.plan(intent, context=_plan_ctx))
                    else:
                        plan = await loop.run_in_executor(None, pl.plan, intent)
    else:
        return JSONResponse({"error": "intent or plan required"}, status_code=400)

    # SoM-Dispatch: einheitlicher Pfad für (a) mehrstufige Intents (som_route)
    # und (b) Groq-Multihop-Versagen ("no plan", SOM_NOPLAN_FALLBACK=1 default).
    # som-planner ist async (antwortet sofort, Ergebnis kommt per Telegram) —
    # blockiert diesen Handler nicht. Reuse build_executor (self-healing gg
    # OpenFang-Agent-respawns).
    if plan is None and intent and (
        som_route or os.environ.get("SOM_NOPLAN_FALLBACK", "1") not in ("0", "false", "False")
    ):
        try:
            from core.capability_targets import build_executor
            ex = build_executor("openfang:som-planner")
            # SYNC call_with_arg (requests.post) → in den Threadpool, sonst
            # blockiert es den Event-Loop + starvt alle anderen Requests
            # (Handler-Starvation, root-caused 2026-06-08).
            import asyncio as _asyncio
            _loop = _asyncio.get_running_loop()
            res = await _loop.run_in_executor(
                None, lambda: ex.call_with_arg(intent, extra_params={"_intent": intent, "_trace_id": trace_id}))
            reply = ""
            if isinstance(res, dict):
                reply = res.get("response") or res.get("final_text") or ""
            _final = (reply.strip() if isinstance(reply, str) and reply.strip()
                      else "An den SoM-Planner übergeben — das Ergebnis kommt per Telegram, sobald der Plan fertig ist.")
            _trace_lite(pe, trace_id, intent, "som-planner", _final)
            return JSONResponse({
                "ok": True, "som": True, "executed": {},
                "final_text": _final, "trace_id": trace_id,
            })
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[multihop] som-planner dispatch failed for intent={intent!r}: {e}")
            # fällt durch auf den "no plan"-Pfad unten

    if plan is None:
        _trace_lite(pe, trace_id, intent, "no-plan", "")
        return JSONResponse({"ok": False, "error": "planner returned no plan", "trace_id": trace_id})

    # Run executor in a thread so FastAPI's main loop is free to handle
    # any nested HTTP calls (brain:GET:/api/X targets recurse into us).
    # NOTE: pe.execute remains sync (ThreadPoolExecutor + thread-mutexes
    # internally) — wrapping it in to_thread is the right call.
    import asyncio as _asyncio
    # E2E-Trace: trace_id auf den Plan setzen, bevor der Executor laeuft — so landet
    # sie im PlanRecorder-Snapshot (plan_executor finally) + stages[] tragen sie.
    try:
        plan.trace_id = trace_id
        plan._stages.append({"stage": "plan", "component": "multihop_execute",
                             "ts": __import__("time").time(), "outcome": "plan_ready"})
    except Exception:  # noqa: BLE001
        pass
    exec_result = await _asyncio.to_thread(pe.execute, plan)
    # MH-5a (Phase 0) — top-level plan_id: the reward-capable correlate.
    # POST /api/multihop/plan/{plan_id}/reward and /api/decisions/reward both
    # key on plan_id; the voice bridge reads data.get("plan_id") — nested-only
    # (plan.plan_id) meant a voice-side reward could never fire.
    out: Dict[str, Any] = {
        "ok": exec_result.get("ok"),
        "trace_id": trace_id,
        "plan_id": plan.plan_id,
        **exec_result,
    }

    # Optional final synthesis — Phase 11.T.4 uses asynthesize() so the
    # synth LLM call doesn't burn a threadpool worker.
    if syn is not None and (intent or plan.intent):
        try:
            asynth = getattr(syn, "asynthesize", None)
            if asynth is not None:
                text = await asynth(
                    intent=intent or plan.intent,
                    plan=plan,
                    executed=exec_result.get("executed", {}),
                    state=exec_result.get("state", {}),
                    custom_prompt=plan.final_synthesis_prompt or None,
                )
            else:
                text = await _asyncio.to_thread(
                    syn.synthesize,
                    intent=intent or plan.intent,
                    plan=plan,
                    executed=exec_result.get("executed", {}),
                    state=exec_result.get("state", {}),
                    custom_prompt=plan.final_synthesis_prompt or None,
                )
            out["final_text"] = text
            # E2E-Trace: die Ausgabe (synthesis) zurueck an den Recorder haengen —
            # bisher wurde final_text NICHT gespeichert (Recon-Befund 2026-06-09).
            try:
                pe.recorder.attach_final(plan.plan_id, text)
            except Exception:  # noqa: BLE001
                pass
        except Exception as e:
            out["synthesis_error"] = f"{type(e).__name__}: {e}"

    return JSONResponse(out)


@router.get("/api/multihop/history")
async def multihop_history(request: Request, limit: int = 20):
    """Phase 6.12 — last N plans (compact). Use /plan/{plan_id} for detail."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"plans": []})
    return JSONResponse({"plans": pe.recorder.list(limit=int(limit))})


# ─── E2E-Trace (2026-06-09): eingabe -> plan -> approval -> execution -> ausgabe ─
@router.get("/api/trace/{trace_id}")
async def trace_get(trace_id: str, request: Request):
    """Volle Nachvollziehbarkeit einer Anfrage: welcher App-Teil (component) hat
    in welcher Stufe (stage) wann (ts) was (outcome) gemacht, plus routed_via +
    final_text (Ausgabe). Funktioniert fuer ALLE Zweige (PlanExecutor + SoM/team
    via Push). 404 wenn unbekannt, nie 500."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"error": "plan_executor not loaded"}, status_code=503)
    try:
        snap = pe.recorder.get_by_trace(trace_id)
    except Exception as e:  # noqa: BLE001
        return JSONResponse({"error": str(e)[:200]}, status_code=500)
    if snap is None:
        return JSONResponse({"error": f"trace '{trace_id}' not found"}, status_code=404)
    # stages chronologisch sortieren (PlanExecutor + SoM-Push koennen verzahnt sein)
    stages = sorted(snap.get("stages", []), key=lambda s: s.get("ts", 0))
    return JSONResponse({
        "trace_id": trace_id,
        "intent": snap.get("intent", ""),
        "routed_via": snap.get("routed_via", ""),
        "plan_id": snap.get("plan_id"),
        "ok": snap.get("ok"),
        "elapsed_s": snap.get("elapsed_s"),
        "stages": stages,
        "executed": snap.get("executed"),
        "final_text": snap.get("final_text", ""),
    })


@router.post("/api/trace/{trace_id}/stage")
async def trace_append_stage(trace_id: str, request: Request):
    """Stage-Push von den detached SoM/som-team-Workern (Phase 2b): {stage,
    component, outcome}. So landen die per-Schritt-Stufen der HEAVY-Kette
    (planner/executor/validator/matrix) unter derselben trace_id im Trace."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"error": "plan_executor not loaded"}, status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    stage = (body.get("stage") or "").strip()
    component = (body.get("component") or "").strip() or "som"
    if not stage:
        return JSONResponse({"error": "stage required"}, status_code=400)
    try:
        pe.recorder.append_stage(trace_id, stage, component, body.get("outcome", ""))
    except Exception as e:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": str(e)[:200]})
    return JSONResponse({"ok": True})


@router.get("/api/trace")
async def trace_list(request: Request, limit: int = 20):
    """Letzte N Traces (trace-zentriert, mit routed_via). Detail via /api/trace/{id}."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"traces": []})
    try:
        with pe.recorder._lock:
            items = [s for s in list(pe.recorder._recent) if s.get("trace_id")][-int(limit):][::-1]
        return JSONResponse({"traces": [
            {"trace_id": s.get("trace_id"), "intent": (s.get("intent") or "")[:120],
             "routed_via": s.get("routed_via"), "ok": s.get("ok"),
             "n_stages": len(s.get("stages", [])), "ts": s.get("ts")}
            for s in items]})
    except Exception as e:  # noqa: BLE001
        return JSONResponse({"traces": [], "error": str(e)[:200]})


@router.get("/api/multihop/plan/{plan_id}")
async def multihop_plan_detail(plan_id: str, request: Request):
    """Phase 6.12 — full plan snapshot incl. per-hop results + state. For
    UI replay."""
    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"error": "plan_executor not loaded"}, status_code=503)
    p = pe.recorder.get(plan_id)
    if p is None:
        return JSONResponse({"error": f"plan '{plan_id}' not found"}, status_code=404)
    return JSONResponse(p)


# ── Phase C — SoM/Team Progress (Push-Modell, Container-Boundary-sicher) ──────
# Die Detached-SoM/Team-Runner POSTen Phasen-Fortschritt hierher (sie sehen den
# Brain via BRAIN_URL/:5000); GET liefert das Live-Dashboard. Kein Mount, kein
# Minibook-Revival nötig — siehe core/som_progress.SomProgressRegistry.
@router.post("/api/som/progress")
async def som_progress_push(request: Request):
    """Runner meldet eine Status-Transition: {run_id, status, intent?, source?}."""
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    run_id = (body.get("run_id") or "").strip()
    status = (body.get("status") or "").strip()
    if not run_id or not status:
        return JSONResponse({"error": "run_id and status required"}, status_code=400)
    try:
        from core.som_progress import get_registry
        get_registry().record(run_id, status,
                              intent=body.get("intent"), source=body.get("source"))
    except Exception as e:  # noqa: BLE001 — Progress darf nie 500'en
        logger.warning(f"[som-progress] record failed: {e}")
        return JSONResponse({"ok": False, "error": str(e)[:200]})
    return JSONResponse({"ok": True})


@router.get("/api/som/runs")
async def som_runs_dashboard(request: Request):
    """Live-Dashboard: laufende + zuletzt fertige SoM/Team-Runs (Push-Registry)."""
    try:
        from core.som_progress import get_registry
        return JSONResponse(get_registry().snapshot())
    except Exception as e:  # noqa: BLE001
        return JSONResponse({"runs": [], "active": [], "done": [], "error": str(e)[:200]})


@router.get("/api/multihop/stream")
async def multihop_stream(request: Request):
    """Phase 6.11 — Server-Sent Events stream. Subscribers get plan/hop
    progress in real time. Events:
      plan_started   payload: full Plan dict
      hop_started    payload: hop spec + rendered_arg
      hop_completed  payload: HopResult preview (incl. kg_hits, validator)
      plan_replanned payload: trigger_step + new_hop_ids
      plan_completed payload: ok + elapsed_s + plan_id
    Connect via JS:  new EventSource('/api/multihop/stream')
    """
    import asyncio as _asyncio
    import json as _json

    pe = getattr(request.app.state, "plan_executor", None)
    if pe is None:
        return JSONResponse({"error": "plan_executor not loaded"}, status_code=503)

    # Bind FastAPI's running loop so background-thread publishes can reach
    # the queue safely. attach_loop() is idempotent.
    loop = _asyncio.get_running_loop()
    if hasattr(pe, "attach_loop"):
        pe.attach_loop(loop)

    q = pe.subscribe()

    async def gen():
        try:
            yield ":connected\n\n"
            while True:
                if await request.is_disconnected():
                    break
                try:
                    event = await _asyncio.wait_for(q.get(), timeout=15)
                except _asyncio.TimeoutError:
                    yield ":keepalive\n\n"
                    continue
                kind = event.get("kind", "message")
                payload = event.get("payload", {})
                yield f"event: {kind}\ndata: {_json.dumps(payload, default=str)}\n\n"
        finally:
            try:
                pe.unsubscribe(q)
            except Exception:
                pass

    from fastapi.responses import StreamingResponse
    return StreamingResponse(gen(), media_type="text/event-stream")


@router.get("/api/self_awareness/state")
async def self_awareness_state(request: Request):
    """One-shot snapshot of Brain's self-awareness layer:
    - substrate (S.1): how many concepts in brain-semantic with self_awareness=True
    - watcher (S.4): tick stats + manifest size
    - aggregated (S.5 input): topics/findings/decisions counts
    - meta_topics (S.5 output): how many cross-session themes exist
    - last self-aware tweet preview from DiscourseEngine
    """
    out: Dict[str, Any] = {"timestamp": time.time()}

    # 1. Substrate count via Qdrant
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is not None:
        try:
            from qdrant_client.http import models as qm
            from core.qdrant_kg import COLLECTIONS
            sem_coll = COLLECTIONS.get("semantic")
            agg_coll = COLLECTIONS.get("aggregated")

            substrate = kg.client.count(
                collection_name=sem_coll,
                count_filter=qm.Filter(must=[
                    qm.FieldCondition(
                        key="self_awareness", match=qm.MatchValue(value=True),
                    ),
                ]),
                exact=True,
            ).count
            out["substrate_concepts"] = substrate

            # 2. Aggregated counts per node-type
            for nt in ("topic", "finding", "decision", "meta_topic"):
                try:
                    c = kg.client.count(
                        collection_name=agg_coll,
                        count_filter=qm.Filter(must=[
                            qm.FieldCondition(
                                key="node_type", match=qm.MatchValue(value=nt),
                            ),
                        ]),
                        exact=True,
                    ).count
                    out[f"aggregated_{nt}_count"] = c
                except Exception:
                    out[f"aggregated_{nt}_count"] = None
        except Exception as e:
            out["kg_error"] = str(e)

    # 3. Watcher stats (S.4)
    saw = getattr(request.app.state, "self_awareness_watcher", None)
    if saw is not None:
        out["watcher"] = saw.stats_dict()
    else:
        out["watcher"] = {"enabled": False}

    # 4. Meta-consolidator stats (S.5)
    dmc = getattr(request.app.state, "discourse_memory_consolidator", None)
    if dmc is not None:
        out["meta_consolidator"] = dmc.stats_dict()
    else:
        out["meta_consolidator"] = {"enabled": False}

    # 5. Last self-aware discourse tweet
    de = getattr(request.app.state, "discourse_engine", None)
    if de is not None:
        s = de.stats_dict()
        out["discourse_engine"] = {
            "running": s.get("running"),
            "ticks": s.get("ticks"),
            "tweets_posted": s.get("tweets_posted"),
            "last_tweet_preview": s.get("last_tweet_preview"),
            "agents_loaded": s.get("agents_loaded"),
        }

    # 6. Manifest path + last seed time
    try:
        from pathlib import Path
        import json as _json
        mf = (Path(__file__).resolve().parent.parent.parent
              / "data" / "self_awareness_manifest.json")
        if mf.exists():
            data = _json.loads(mf.read_text(encoding="utf-8"))
            out["manifest"] = {
                "path": str(mf),
                "last_full_seed_at": data.get("last_full_seed_at"),
                "last_checked_at": data.get("last_checked_at"),
                "source_count": len(data.get("sources") or {}),
            }
    except Exception as e:
        out["manifest_error"] = str(e)

    return JSONResponse(out)


# ─────────────────────────────────────────────────────────────────────
# Phase R+ — Three-Mode Discourse (Intent + Response triggers)
# ─────────────────────────────────────────────────────────────────────


@router.post("/api/discourse/intent")
async def discourse_intent(request: Request):
    """Trigger an Intent-Mode discourse round.

    Body: {"message": "...", "context": {...}, "auto_dispatch": true}

    Runs all 26 phi3-clones in parallel against the user intent. Returns
    the aggregator's decision JSON. If `auto_dispatch=true` (default) AND
    the decision confidence ≥ threshold, also fires a real OpenFang call
    to the chosen primary agent and includes its answer in the response.

    Returns:
        {
          "decision": {primary, supporting, risks, confidence, reasoning},
          "tweet_count": int,
          "high_confidence": bool,
          "dispatched": {agent_id, agent_name, response, ...} | null,
        }
    """
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"},
                            status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    msg = (body.get("message") or "").strip()
    if not msg:
        return JSONResponse({"error": "message required"}, status_code=400)
    auto = bool(body.get("auto_dispatch", True))
    ctx = body.get("context") or {}

    # Phase 11.U.B — tick_intent is synchronous and can take 30-60s
    # (parallel agent calls + groq aggregator). Offload to thread so the
    # FastAPI event loop keeps serving /api/health, SSE streams, etc.
    discourse = await asyncio.to_thread(de.tick_intent, msg, ctx)

    out = {
        "ok":              bool(discourse.get("ok", True)),
        "decision":        discourse.get("decision") or {},
        "tweets":          discourse.get("tweets") or [],
        "tweet_count":     discourse.get("tweet_count", 0),
        "high_confidence": bool(discourse.get("high_confidence")),
        "dispatched":      None,
        # Phase 1 capability routing fields — None when no match
        "capability":      discourse.get("capability"),
        "matched_pattern": discourse.get("matched_pattern"),
        "agents_targeted": discourse.get("agents_targeted"),
        "agents_total":    discourse.get("agents_total"),
        # Phase 1.5 direct-execution fields — None for normal broadcast path
        "is_direct":        discourse.get("is_direct"),
        "direct_target":    discourse.get("direct_target"),
        "direct_elapsed_s": discourse.get("direct_elapsed_s"),
        "direct_error":     discourse.get("direct_error"),
        "result":           discourse.get("result"),
    }

    # Confidence-aware dispatch (R+.8) — Mode A in plan
    # Phase 11.U.B — also off-thread; the OpenFang call has timeout=600s
    if auto and discourse.get("high_confidence"):
        decision = discourse.get("decision") or {}
        primary = decision.get("primary")
        if primary:
            dispatched = await asyncio.to_thread(
                _dispatch_to_openfang, primary, msg, decision,
            )
            out["dispatched"] = dispatched

    return JSONResponse(out)


def _dispatch_to_openfang(
    agent_name: str, task: str, decision: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """Find the agent in OpenFang by name (resolve UUID), POST the task,
    return the agent's response payload."""
    import os
    import requests
    of_url = os.environ.get("OPENFANG_URL", "http://127.0.0.1:4200").rstrip("/")
    try:
        r = requests.get(f"{of_url}/api/agents", timeout=10)
        agents = r.json() if r.ok else []
        # Prefer non-phi3 (Sonnet) for actual execution
        target = None
        norm = (agent_name or "").lower().rstrip("-phi3")
        for a in agents:
            n = (a.get("name") or "").lower()
            if n == norm or n == agent_name.lower():
                target = a
                break
        if target is None:
            return {"error": f"agent '{agent_name}' not found in OpenFang"}
        agent_id = target.get("id")
        # Compose context: include supporting + risks for the agent
        supporting = decision.get("supporting") or []
        risks = decision.get("risks") or []
        context_parts = [f"Task: {task[:1500]}"]
        if supporting:
            context_parts.append(f"Supporting agents flagged: {', '.join(supporting)}")
        if risks:
            context_parts.append(
                "Risks raised by other agents:\n" +
                "\n".join(f"  - {r}" for r in risks[:5])
            )
        composed = "\n\n".join(context_parts)
        r = requests.post(
            f"{of_url}/api/agents/{agent_id}/message",
            json={"message": composed[:60000], "sender_name": "Brain"},
            timeout=600,
        )
        if not r.ok:
            return {"error": f"openfang HTTP {r.status_code}",
                    "body": r.text[:300]}
        return {
            "agent_id":  agent_id,
            "agent_name": target.get("name"),
            "response":  ((r.json() or {}).get("response") or "")[:5000],
            "input_tokens":  (r.json() or {}).get("input_tokens"),
            "output_tokens": (r.json() or {}).get("output_tokens"),
            "iterations":    (r.json() or {}).get("iterations"),
            "cost_usd":      (r.json() or {}).get("cost_usd"),
        }
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}"}


@router.post("/api/discourse/response")
async def discourse_response(request: Request):
    """Manually queue a Brain-response for the next Mode-3 tick.

    Body: {"response_text": "..."}
    Returns: {"queued": true, "queue_depth": N}
    """
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"},
                            status_code=503)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    text = (body.get("response_text") or "").strip()
    if not text:
        return JSONResponse({"error": "response_text required"}, status_code=400)
    de.queue_response(text, body.get("context") or {})
    return JSONResponse({
        "queued": True,
        "queue_depth": len(getattr(de, "_response_queue", []) or []),
    })


@router.post("/api/discourse/response_tick_now")
async def discourse_response_tick_now(request: Request):
    """Force one Mode-3 tick. Pulls oldest from queue (if any) and asks
    3-5 random agents to assess."""
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"},
                            status_code=503)
    return JSONResponse(de.tick_response())


@router.get("/api/discourse/intent_decisions")
async def discourse_intent_decisions(request: Request, limit: int = 10):
    """Last N intent-mode decisions (in-memory ring buffer)."""
    de = getattr(request.app.state, "discourse_engine", None)
    if de is None:
        return JSONResponse({"error": "discourse_engine not running"},
                            status_code=503)
    return JSONResponse({
        "count": len(de.intent_decisions(limit)) if hasattr(de, "intent_decisions") else 0,
        "decisions": de.intent_decisions(limit) if hasattr(de, "intent_decisions") else [],
    })


@router.get("/api/kg/mcmp_stats")
async def kg_mcmp_stats(request: Request):
    """Stats about the MCMP gardener (random walker + pruner)."""
    g = getattr(request.app.state, "mcmp_gardener", None)
    if g is None:
        return JSONResponse(
            {"enabled": False, "message": "mcmp_gardener not running"},
        )
    return JSONResponse({
        "enabled": True,
        "stats": convert_numpy(dict(g.stats)),
        "config": {
            "tick_interval_s": float(__import__("core.mcmp_gardener", fromlist=["TICK_INTERVAL_S"]).TICK_INTERVAL_S),
        },
        "timestamp": time.time(),
    })


# ===================================================================
# Group 7 — Goals / Evolution / CTM / Cognitive Status
# ===================================================================

@router.get("/api/brain/goals")
async def brain_goals(request: Request):
    """Current goals — graceful fallback."""
    return JSONResponse({
        "goals": [],
        "enabled": False,
        "message": "goal system not connected to unified brain",
        "timestamp": time.time(),
    })


@router.post("/api/brain/goals/add")
async def brain_goals_add(request: Request):
    """Add a goal — 503 until wired."""
    return JSONResponse(
        {"error": "goal system not initialized", "timestamp": time.time()},
        status_code=503,
    )


@router.post("/api/brain/goals/{goal_id}/complete")
async def brain_goals_complete(goal_id: str, request: Request):
    """Mark goal complete — 503 until wired."""
    return JSONResponse(
        {"error": "goal system not initialized", "goal_id": goal_id, "timestamp": time.time()},
        status_code=503,
    )


@router.post("/api/brain/goals/{goal_id}/fail")
async def brain_goals_fail(goal_id: str, request: Request):
    """Mark goal failed — 503 until wired."""
    return JSONResponse(
        {"error": "goal system not initialized", "goal_id": goal_id, "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/brain/evolution")
async def brain_evolution(request: Request):
    """Evolution state — graceful fallback."""
    return JSONResponse({
        "evolution": None,
        "enabled": False,
        "message": "evolution system not connected to unified brain",
        "timestamp": time.time(),
    })


@router.post("/api/brain/evolution/evolve")
async def brain_evolution_evolve(request: Request):
    """Trigger evolution step — 503 until wired."""
    return JSONResponse(
        {"error": "evolution system not initialized", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/brain/ctm_health")
async def ctm_health(request: Request):
    """CTM health — graceful fallback."""
    return JSONResponse({
        "ctm_health": None,
        "enabled": False,
        "message": "CTM not connected to unified brain",
        "timestamp": time.time(),
    })


@router.get("/api/brain/cognitive_status")
async def cognitive_status(request: Request):
    """Cognitive status summary — graceful fallback."""
    return JSONResponse({
        "cognitive_status": None,
        "enabled": False,
        "message": "cognitive status not connected to unified brain",
        "timestamp": time.time(),
    })


# ===================================================================
# Group 8 — Causal / Meta / Federated / Advanced Learning
# ===================================================================

@router.get("/api/causal/status")
async def causal_status(request: Request):
    """Causal reasoning status — not yet wired."""
    return JSONResponse(
        {"error": "causal reasoning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/causal/graph")
async def causal_graph(request: Request):
    """Causal graph — not yet wired."""
    return JSONResponse(
        {"error": "causal reasoning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.post("/api/causal/analyze")
async def causal_analyze(request: Request):
    """Causal analysis — not yet wired."""
    return JSONResponse(
        {"error": "causal reasoning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/meta/status")
async def meta_status(request: Request):
    """Meta-learning status — not yet wired."""
    return JSONResponse(
        {"error": "meta-learning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.post("/api/meta/adapt")
async def meta_adapt(request: Request):
    """Meta-learning adapt — not yet wired."""
    return JSONResponse(
        {"error": "meta-learning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/federated/status")
async def federated_status(request: Request):
    """Federated learning status — not yet wired."""
    return JSONResponse(
        {"error": "federated learning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/federated/nodes")
async def federated_nodes(request: Request):
    """Federated learning nodes — not yet wired."""
    return JSONResponse(
        {"error": "federated learning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/federated/rounds")
async def federated_rounds(request: Request):
    """Federated learning rounds — not yet wired."""
    return JSONResponse(
        {"error": "federated learning not available", "timestamp": time.time()},
        status_code=503,
    )


@router.get("/api/advanced_learning/health")
async def advanced_learning_health(request: Request):
    """Advanced learning health — not yet wired."""
    return JSONResponse(
        {"error": "advanced learning not available", "timestamp": time.time()},
        status_code=503,
    )


# ===================================================================
# Group 9 — Conversation Monitoring & Simulation
# ===================================================================

@router.get("/api/conversation/active")
async def conversation_active(request: Request):
    """Active conversations from LiveBrainMonitor."""
    lm = request.app.state.live_monitor
    if lm is None:
        return JSONResponse({
            "conversations": [],
            "message": "live_monitor not initialized",
            "timestamp": time.time(),
        })
    try:
        convos = lm.get_active_conversations()
        return JSONResponse({
            "conversations": convert_numpy(convos),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "conversations": [],
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.get("/api/conversation/history")
async def conversation_history(request: Request):
    """Conversation history."""
    lm = request.app.state.live_monitor
    if lm is None:
        return JSONResponse({
            "history": [],
            "message": "live_monitor not initialized",
            "timestamp": time.time(),
        })
    try:
        history = lm.get_conversation_history()
        return JSONResponse({
            "history": convert_numpy(history),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "history": [],
            "error": str(exc),
            "timestamp": time.time(),
        })


@router.post("/api/simulate/conversation")
async def simulate_conversation(request: Request):
    """Simulate a conversation for testing — 503 until wired."""
    return JSONResponse(
        {"error": "simulation not available", "timestamp": time.time()},
        status_code=503,
    )


# ===================================================================
# Group 10 — Predict Path
# ===================================================================

@router.post("/api/predict/path")
async def predict_path(request: Request):
    """Predict conversation path using ConversationPathPlanner."""
    pp = request.app.state.path_planner
    if pp is None:
        return JSONResponse({
            "path": None,
            "message": "path_planner not initialized",
            "timestamp": time.time(),
        })
    try:
        body = await request.json()
        task = body.get("task", "")
        if not task:
            return JSONResponse({"error": "task is required"}, status_code=400)
        result = pp.predict_path(task)
        return JSONResponse({
            "path": convert_numpy(result),
            "timestamp": time.time(),
        })
    except Exception as exc:
        return JSONResponse({
            "path": None,
            "error": str(exc),
            "timestamp": time.time(),
        })


# ===================================================================
# Phase 10 — Self-Reflective Decision Loop endpoints
# ===================================================================

@router.get("/api/decisions/recall")
async def decisions_recall(request: Request, q: str = "", k: int = 5) -> JSONResponse:
    """Recall past decisions for an intent. ?q=<intent>&k=5"""
    kg = getattr(request.app.state, "qdrant_kg", None)
    if not q.strip() or kg is None:
        return JSONResponse({"hits": [], "count": 0})
    try:
        from core import decision_recall
        hits = decision_recall.recall(q, kg, k=k)
        return JSONResponse({
            "hits": hits, "count": len(hits), "query": q, "timestamp": time.time(),
        })
    except Exception as e:
        return JSONResponse({"hits": [], "error": str(e)}, status_code=500)


@router.post("/api/decisions/reward")
async def decisions_reward(request: Request) -> JSONResponse:
    """Attach an explicit reward to a past decision and propagate to self-model.
    Body: {plan_id, reward: -1..1, comment?}"""
    body = await request.json()
    plan_id = (body or {}).get("plan_id", "")
    reward = float((body or {}).get("reward", 0.0))
    if not plan_id:
        return JSONResponse({"ok": False, "error": "plan_id required"}, status_code=400)
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse({"ok": False, "error": "kg unavailable"}, status_code=503)
    try:
        # Update decision_record reward field
        from core.qdrant_kg import _point_id
        ext_id = f"decision::{plan_id}"
        pid = _point_id(ext_id)
        recs = kg.client.retrieve(
            collection_name="brain-decisions", ids=[pid],
            with_payload=True, with_vectors=False,
        )
        if not recs:
            return JSONResponse({"ok": False, "error": "decision not found"}, status_code=404)
        existing = dict(recs[0].payload or {})
        existing["reward"] = reward
        existing["reward_comment"] = (body or {}).get("comment", "")[:300]
        # Re-upsert via plain payload-set (no re-embed needed)
        from qdrant_client.http.models import PointStruct  # type: ignore
        kg.client.set_payload(
            collection_name="brain-decisions",
            payload=existing, points=[pid],
        )
        # Propagate to self-model: update each capability used
        from core import decision_self_prior
        intent = existing.get("intent", "")
        for cap in (existing.get("capability_chain") or []):
            decision_self_prior.update(
                intent_text=intent, capability=cap,
                success=(reward > 0), reward=reward,
                plan_id=plan_id, kg=kg,
            )
        return JSONResponse({
            "ok": True, "plan_id": plan_id, "reward": reward,
            "capabilities_updated": len(existing.get("capability_chain") or []),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/self/prior")
async def self_prior(request: Request, q: str = "", k: int = 8) -> JSONResponse:
    """Get capability-confidence prior for an intent. ?q=<intent>&k=8"""
    kg = getattr(request.app.state, "qdrant_kg", None)
    if not q.strip() or kg is None:
        return JSONResponse({"capabilities": [], "best_capability": None})
    try:
        from core import decision_self_prior
        return JSONResponse(decision_self_prior.prior(q, kg, k=k))
    except Exception as e:
        return JSONResponse({"capabilities": [], "error": str(e)}, status_code=500)


@router.get("/api/self/snapshot")
async def self_snapshot(request: Request, limit: int = 200) -> JSONResponse:
    """Full self-model dump (sorted by n_observations desc)."""
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse({"traits": [], "count": 0})
    try:
        from core import decision_self_prior
        traits = decision_self_prior.snapshot(kg, limit=limit)
        return JSONResponse({
            "traits": traits, "count": len(traits), "timestamp": time.time(),
        })
    except Exception as e:
        return JSONResponse({"traits": [], "error": str(e)}, status_code=500)


@router.post("/api/critic/preview")
async def critic_preview(request: Request) -> JSONResponse:
    """Run plan_critic on an arbitrary intent without executing.
    Body: {intent} -> generates a quick plan via planner then critiques it."""
    body = await request.json()
    intent = (body or {}).get("intent", "")
    if not intent.strip():
        return JSONResponse({"ok": False, "error": "intent required"}, status_code=400)
    pe = getattr(request.app.state, "plan_executor", None)
    planner = getattr(request.app.state, "multihop_planner", None) or \
              getattr(request.app.state, "plan_generator", None)
    if pe is None or planner is None:
        return JSONResponse({"ok": False, "error": "planner unavailable"}, status_code=503)
    try:
        plan = planner.plan(intent) if hasattr(planner, "plan") else None
        if plan is None:
            return JSONResponse({"ok": False, "error": "plan generation failed"}, status_code=500)
        from core import plan_critic
        verdict = plan_critic.critique(plan, intent, pe.dispatcher)
        return JSONResponse({
            "ok": True, "intent": intent,
            "plan_id": getattr(plan, "plan_id", None),
            "verdict": verdict,
            "plan": plan.to_dict() if hasattr(plan, "to_dict") else None,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ===================================================================
# Phase 11.C — Event Mapping (Event → Agent → MCP-Tool)
# ===================================================================

@router.get("/api/events/mapping")
async def events_mapping(request: Request) -> JSONResponse:
    """Join all events with their claiming agents and MCP-tools.

    Phase 11.D — uses two new sources:
      - AgentYamlRegistry (configs/agents/*.yaml) for explicit event->agent
        mappings. Auto-seeded from namespace-defaults on first call.
      - McpDiscovery (data/mcp_tools_cache.json) for live tool inventory.

    Returns per event: {event_id, namespace, agent, agent_source,
                        tool, tool_description, tool_args, coverage}
    Plus: per-namespace stats, full agent list, full tool inventory.
    """
    import os, re, json
    kg = getattr(request.app.state, "qdrant_kg", None)
    if kg is None:
        return JSONResponse({"events": [], "agents": [], "tools": [],
                             "namespaces": {}, "error": "kg unavailable"})

    try:
        from core.agent_yaml_registry import get_registry
        from core.mcp_discovery import get_discovery
        registry = get_registry()
        discovery = get_discovery()
        registry.reload_if_changed()
        # 1. Pull all 137 events from procedural collection
        events = []
        try:
            recs, _ = kg.client.scroll(
                collection_name="brain-procedural",
                limit=400,
                with_payload=True, with_vectors=False,
            )
            for r in recs or []:
                p = dict(r.payload or {})
                if p.get("node_type") == "event":
                    eid = p.get("event_id") or p.get("title") or ""
                    if not eid:
                        continue
                    ns = eid.split(".", 1)[0] if "." in eid else "(other)"
                    events.append({
                        "event_id": eid,
                        "namespace": ns,
                        "title": p.get("title", eid),
                        "description": (p.get("description") or "")[:200],
                    })
        except Exception as e:
            return JSONResponse({"events": [], "error": f"kg scroll: {e}"},
                                status_code=500)

        # 2. Read OpenFang agent manifests from filesystem
        agents = []
        try:
            agent_dir = "C:/Users/User/Desktop/Vibemind_V1/vibemind-os/openfang/agents"
            if os.path.isdir(agent_dir):
                for sub in sorted(os.listdir(agent_dir)):
                    sub_path = os.path.join(agent_dir, sub)
                    toml_path = os.path.join(sub_path, "agent.toml")
                    if not os.path.isfile(toml_path):
                        continue
                    txt = open(toml_path, encoding="utf-8").read()
                    name_m = re.search(r'^name\s*=\s*"([^"]+)"', txt, re.M)
                    desc_m = re.search(r'^description\s*=\s*"([^"]+)"', txt, re.M)
                    tags_m = re.search(r'^tags\s*=\s*\[([^\]]+)\]', txt, re.M)
                    mcp_m = re.search(
                        r'\[mcp_allowed\]\s*\nservers\s*=\s*\[([^\]]+)\]', txt
                    )
                    model_m = re.search(r'^model\s*=\s*"([^"]+)"', txt, re.M)
                    tags = []
                    if tags_m:
                        tags = [t.strip().strip('"') for t in tags_m.group(1).split(",")]
                    mcp_servers = []
                    if mcp_m:
                        mcp_servers = [t.strip().strip('"')
                                       for t in mcp_m.group(1).split(",")]
                    spaces_in_tags = [t for t in tags if t.startswith("space:")]
                    agents.append({
                        "name": name_m.group(1) if name_m else sub,
                        "description": desc_m.group(1) if desc_m else "",
                        "tags": tags,
                        "spaces": [s.replace("space:", "") for s in spaces_in_tags],
                        "mcp_servers": mcp_servers,
                        "model": model_m.group(1) if model_m else "",
                    })
        except Exception as e:
            agents = []
            err_agents = str(e)

        # 3. Compute namespace -> default agent (legacy heuristic) for fallback.
        # Events use singular namespaces (bubble.create) but agents use plural
        # space tags (space:bubbles). Strip trailing 's' on space tag for match.
        def _normalize_ns(ns: str) -> list:
            """Return possible normalisations: ['bubbles', 'bubble'] etc."""
            out = [ns]
            if ns.endswith("s") and len(ns) > 2:
                out.append(ns[:-1])
            return out
        ns_to_agent_default = {}
        for a in agents:
            for s in a.get("spaces", []):
                for variant in _normalize_ns(s):
                    ns_to_agent_default.setdefault(variant, []).append(a["name"])
            n = a["name"].replace("brain-", "").replace("-phi3", "")
            if n:
                for variant in _normalize_ns(n):
                    if variant not in ns_to_agent_default:
                        ns_to_agent_default.setdefault(variant, []).append(a["name"])

        # 4. Auto-seed registry from namespace defaults (idempotent)
        events_by_ns: Dict[str, list] = {}
        for ev in events:
            events_by_ns.setdefault(ev["namespace"], []).append(ev["event_id"])
        # Pick first agent per namespace as the default for seeding
        ns_to_agent_seed = {
            ns: agents_list[0] for ns, agents_list in ns_to_agent_default.items()
            if agents_list
        }
        if registry.stats_dict().get("events_total", 0) == 0:
            registry.auto_seed(ns_to_agent_seed, events_by_ns)

        # 5. Build per-event mapping using registry (explicit) + discovery (tool)
        mapped = []
        agent_lookup = {a["name"]: a for a in agents}
        for ev in events:
            ns = ev["namespace"]
            # 5a. Resolve agent: explicit YAML claim > namespace default
            assigned_agent = registry.get_event_agent(ev["event_id"])
            agent_source = "yaml" if assigned_agent else None
            if not assigned_agent:
                claims = ns_to_agent_default.get(ns, [])
                if claims:
                    assigned_agent = claims[0]
                    agent_source = "namespace_default"

            # 5b. Resolve tool via live discovery using agent's mcp_allowed
            tool_info = None
            if assigned_agent and assigned_agent in agent_lookup:
                mcp_srv = agent_lookup[assigned_agent].get("mcp_servers", [])
                tool_info = discovery.find_tool_for_event(ev["event_id"], mcp_srv)

            tool = ""
            tool_desc = ""
            tool_args = {}
            if tool_info:
                tool = f"{tool_info['server']}/{tool_info['tool']}"
                tool_desc = tool_info.get("description", "")
                tool_args = tool_info.get("input_schema", {})

            coverage = "full" if (assigned_agent and tool) else (
                "partial" if assigned_agent else "none"
            )
            mapped.append({
                "event_id": ev["event_id"],
                "namespace": ns,
                "title": ev["title"],
                "description": ev["description"],
                "agents_claiming": ns_to_agent_default.get(ns, []),
                "agent": assigned_agent or "",
                "agent_source": agent_source or "",
                "tool": tool,
                "tool_description": tool_desc,
                "tool_args": tool_args,
                "coverage": coverage,
            })

        # 6. Aggregate stats
        ns_stats = {}
        for m in mapped:
            ns = m["namespace"]
            ns_stats.setdefault(ns, {"total": 0, "full": 0, "partial": 0, "none": 0})
            ns_stats[ns]["total"] += 1
            ns_stats[ns][m["coverage"]] += 1

        return JSONResponse({
            "ok": True,
            "events_count": len(mapped),
            "agents_count": len(agents),
            "events": mapped,
            "agents": agents,
            "mcp_servers": discovery.list_servers(),
            "namespaces": ns_stats,
            "registry_stats": registry.stats_dict(),
            "discovery_stats": discovery.stats_dict(),
            "timestamp": time.time(),
        })

    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.post("/api/events/assign")
async def events_assign(request: Request) -> JSONResponse:
    """Move an event from one agent to another.
    Body: {event_id, to_agent, from_agent? (auto-resolved if missing)}
    """
    try:
        from core.agent_yaml_registry import get_registry
        body = await request.json()
        event_id = (body or {}).get("event_id", "").strip()
        to_agent = (body or {}).get("to_agent", "").strip()
        from_agent = (body or {}).get("from_agent", "").strip()
        if not event_id or not to_agent:
            return JSONResponse(
                {"ok": False, "error": "event_id and to_agent required"},
                status_code=400,
            )
        registry = get_registry()
        if not from_agent:
            from_agent = registry.get_event_agent(event_id) or ""
        result = registry.move_event(event_id, from_agent, to_agent)
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.post("/api/events/unassign")
async def events_unassign(request: Request) -> JSONResponse:
    """Remove an event's claim from an agent.
    Body: {event_id, from_agent}
    """
    try:
        from core.agent_yaml_registry import get_registry
        body = await request.json()
        event_id = (body or {}).get("event_id", "").strip()
        from_agent = (body or {}).get("from_agent", "").strip()
        if not event_id or not from_agent:
            return JSONResponse(
                {"ok": False, "error": "event_id and from_agent required"},
                status_code=400,
            )
        registry = get_registry()
        return JSONResponse(registry.remove_event(event_id, from_agent))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/agents/yaml/list")
async def agents_yaml_list(request: Request) -> JSONResponse:
    """List all agent-YAML manifests with their claimed events."""
    try:
        from core.agent_yaml_registry import get_registry
        registry = get_registry()
        registry.reload_if_changed()
        return JSONResponse({
            "ok": True,
            "agents": registry.list_agents(),
            "stats": registry.stats_dict(),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/agents/{agent_name}/tools")
async def agent_tools(request: Request, agent_name: str) -> JSONResponse:
    """Detailed tool list for one agent: all tools from all servers in
    that agent's mcp_allowed list. Used by UI to show 'alternative tools
    available' when default tool-resolution doesn't pick the right one.
    """
    import os, re
    try:
        from core.mcp_discovery import get_discovery
        # Read agent.toml to get mcp_allowed
        agent_dir = "C:/Users/User/Desktop/Vibemind_V1/vibemind-os/openfang/agents/" + agent_name
        toml_path = os.path.join(agent_dir, "agent.toml")
        if not os.path.isfile(toml_path):
            return JSONResponse({"ok": False, "error": "agent not found"},
                                status_code=404)
        txt = open(toml_path, encoding="utf-8").read()
        mcp_m = re.search(r'\[mcp_allowed\]\s*\nservers\s*=\s*\[([^\]]+)\]', txt)
        mcp_servers = []
        if mcp_m:
            mcp_servers = [t.strip().strip('"')
                           for t in mcp_m.group(1).split(",")]
        discovery = get_discovery()
        tools_by_server = {}
        for srv in mcp_servers:
            tools_by_server[srv] = discovery.list_tools(srv)
        return JSONResponse({
            "ok": True,
            "agent": agent_name,
            "mcp_servers": mcp_servers,
            "tools_by_server": tools_by_server,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.post("/api/mcp/discover")
async def mcp_discover(request: Request) -> JSONResponse:
    """Trigger fresh MCP-tool discovery (parallel spawn all servers).
    Takes 5-15s depending on server count. Updates cache."""
    try:
        from core.mcp_discovery import get_discovery
        discovery = get_discovery()
        result = discovery.discover_all()
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/events/mapping.xlsx")
async def events_mapping_xlsx(request: Request):
    """Phase 11.E — Excel export of the full event-mapping inventory.

    Returns a 4-sheet workbook:
      Sheet 1 'Events'   — all 137 events × namespace × agent × tool × coverage
      Sheet 2 'Agents'   — all OpenFang agents (name, model, tags, mcp_servers)
      Sheet 3 'Tools'    — all discovered MCP tools (server, name, description, args)
      Sheet 4 'Coverage' — pivot summary per namespace
    """
    from fastapi.responses import Response
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        return JSONResponse({"ok": False, "error": "openpyxl not installed"},
                            status_code=500)
    try:
        from core.agent_yaml_registry import get_registry
        from core.mcp_discovery import get_discovery
        registry = get_registry()
        discovery = get_discovery()
        registry.reload_if_changed()

        # Re-fetch the same data the JSON endpoint produces
        # (simplified: we call ourselves internally via direct python calls)
        kg = getattr(request.app.state, "qdrant_kg", None)
        if kg is None:
            return JSONResponse({"ok": False, "error": "kg unavailable"},
                                status_code=503)

        # 1. Events
        events = []
        recs, _ = kg.client.scroll(
            collection_name="brain-procedural", limit=400,
            with_payload=True, with_vectors=False,
        )
        for r in recs or []:
            p = dict(r.payload or {})
            if p.get("node_type") == "event":
                eid = p.get("event_id") or p.get("title") or ""
                if not eid:
                    continue
                ns = eid.split(".", 1)[0] if "." in eid else "(other)"
                events.append({
                    "event_id": eid, "namespace": ns,
                    "title": p.get("title", eid),
                    "description": (p.get("description") or "")[:200],
                })

        # 2. Agents from OpenFang manifests
        import os, re
        agents = []
        agent_dir = "C:/Users/User/Desktop/Vibemind_V1/vibemind-os/openfang/agents"
        if os.path.isdir(agent_dir):
            for sub in sorted(os.listdir(agent_dir)):
                toml_path = os.path.join(agent_dir, sub, "agent.toml")
                if not os.path.isfile(toml_path):
                    continue
                txt = open(toml_path, encoding="utf-8").read()
                name_m = re.search(r'^name\s*=\s*"([^"]+)"', txt, re.M)
                desc_m = re.search(r'^description\s*=\s*"([^"]+)"', txt, re.M)
                tags_m = re.search(r'^tags\s*=\s*\[([^\]]+)\]', txt, re.M)
                mcp_m = re.search(
                    r'\[mcp_allowed\]\s*\nservers\s*=\s*\[([^\]]+)\]', txt
                )
                model_m = re.search(r'^model\s*=\s*"([^"]+)"', txt, re.M)
                tags = []
                if tags_m:
                    tags = [t.strip().strip('"') for t in tags_m.group(1).split(",")]
                mcp_servers = []
                if mcp_m:
                    mcp_servers = [t.strip().strip('"')
                                   for t in mcp_m.group(1).split(",")]
                agents.append({
                    "name": name_m.group(1) if name_m else sub,
                    "description": desc_m.group(1) if desc_m else "",
                    "tags": tags,
                    "mcp_servers": mcp_servers,
                    "model": model_m.group(1) if model_m else "",
                })

        agent_lookup = {a["name"]: a for a in agents}

        # Build agent-resolution like events_mapping does
        def _normalize_ns(ns: str) -> list:
            out = [ns]
            if ns.endswith("s") and len(ns) > 2:
                out.append(ns[:-1])
            return out
        ns_to_agent_default = {}
        for a in agents:
            for tag in a.get("tags", []):
                if tag.startswith("space:"):
                    s = tag.replace("space:", "")
                    for variant in _normalize_ns(s):
                        ns_to_agent_default.setdefault(variant, []).append(a["name"])
            n = a["name"].replace("brain-", "").replace("-phi3", "")
            if n:
                for variant in _normalize_ns(n):
                    if variant not in ns_to_agent_default:
                        ns_to_agent_default.setdefault(variant, []).append(a["name"])

        # ── Build the workbook ──────────────────────────────────────
        wb = openpyxl.Workbook()
        # Styling helpers
        HEADER_FILL = PatternFill(start_color="0F1923", end_color="0F1923", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="4FC3F7", size=11)
        FULL_FILL = PatternFill(start_color="1B4D2A", end_color="1B4D2A", fill_type="solid")
        PARTIAL_FILL = PatternFill(start_color="4D3D14", end_color="4D3D14", fill_type="solid")
        NONE_FILL = PatternFill(start_color="4D1A1A", end_color="4D1A1A", fill_type="solid")
        thin = Side(border_style="thin", color="2A3F54")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _style_header_row(ws, row=1):
            for cell in ws[row]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER
            ws.freeze_panes = "A2"

        # ─── Sheet 1: Events ─────────────────────────────────────────
        ws_events = wb.active
        ws_events.title = "Events"
        ws_events.append([
            "Namespace", "Event ID", "Event Description",
            "Agent", "Agent Source", "MCP Server", "MCP Tool",
            "Tool Description", "Tool Args (JSON)", "Coverage",
        ])
        _style_header_row(ws_events)
        col_widths = [14, 32, 50, 22, 18, 18, 28, 60, 40, 12]
        for i, w in enumerate(col_widths, start=1):
            ws_events.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        full_count = partial_count = none_count = 0
        for ev in sorted(events, key=lambda e: (e["namespace"], e["event_id"])):
            assigned_agent = registry.get_event_agent(ev["event_id"])
            agent_source = "yaml" if assigned_agent else None
            if not assigned_agent:
                claims = ns_to_agent_default.get(ev["namespace"], [])
                if claims:
                    assigned_agent = claims[0]
                    agent_source = "namespace_default"

            tool_info = None
            if assigned_agent and assigned_agent in agent_lookup:
                mcp_srv = agent_lookup[assigned_agent].get("mcp_servers", [])
                tool_info = discovery.find_tool_for_event(
                    ev["event_id"], mcp_srv,
                )

            mcp_server = tool_info["server"] if tool_info else ""
            tool_name = tool_info["tool"] if tool_info else ""
            tool_desc = (tool_info or {}).get("description", "")
            tool_args = (tool_info or {}).get("input_schema", {})
            try:
                tool_args_str = json.dumps(tool_args, ensure_ascii=False)[:500]
            except Exception:
                tool_args_str = ""

            coverage = "full" if (assigned_agent and tool_name) else (
                "partial" if assigned_agent else "none"
            )
            if coverage == "full":
                full_count += 1
            elif coverage == "partial":
                partial_count += 1
            else:
                none_count += 1

            row = [
                ev["namespace"], ev["event_id"], ev["description"],
                assigned_agent or "", agent_source or "",
                mcp_server, tool_name, tool_desc[:300], tool_args_str,
                coverage,
            ]
            ws_events.append(row)
            # Color the coverage cell
            row_idx = ws_events.max_row
            cov_cell = ws_events.cell(row=row_idx, column=10)
            cov_cell.fill = (
                FULL_FILL if coverage == "full" else
                PARTIAL_FILL if coverage == "partial" else NONE_FILL
            )
            cov_cell.font = Font(bold=True, color="FFFFFF")
            cov_cell.alignment = Alignment(horizontal="center")

        ws_events.auto_filter.ref = ws_events.dimensions

        # ─── Sheet 2: Agents ─────────────────────────────────────────
        ws_agents = wb.create_sheet("Agents")
        ws_agents.append([
            "Agent Name", "Description", "Model",
            "Tags", "MCP Servers Allowed",
            "Events Claimed (from YAML)", "Events Count",
        ])
        _style_header_row(ws_agents)
        for i, w in enumerate([28, 50, 30, 35, 35, 60, 12], start=1):
            ws_agents.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w
        for a in sorted(agents, key=lambda x: x["name"]):
            claimed = registry.get_agent_events(a["name"])
            ws_agents.append([
                a["name"], a.get("description", ""), a.get("model", ""),
                ", ".join(a.get("tags") or []),
                ", ".join(a.get("mcp_servers") or []),
                ", ".join(claimed),
                len(claimed),
            ])
        ws_agents.auto_filter.ref = ws_agents.dimensions

        # ─── Sheet 3: Tools ──────────────────────────────────────────
        ws_tools = wb.create_sheet("Tools")
        ws_tools.append([
            "MCP Server", "Tool Name", "Description", "Args Schema (JSON)",
        ])
        _style_header_row(ws_tools)
        for i, w in enumerate([22, 32, 70, 50], start=1):
            ws_tools.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w
        all_tools = discovery.all_tools_flat()
        for t in sorted(all_tools, key=lambda x: (x.get("server", ""), x.get("name", ""))):
            try:
                args_str = json.dumps(t.get("input_schema") or {}, ensure_ascii=False)[:500]
            except Exception:
                args_str = ""
            ws_tools.append([
                t.get("server", ""), t.get("name", ""),
                (t.get("description") or "")[:400],
                args_str,
            ])
        ws_tools.auto_filter.ref = ws_tools.dimensions

        # ─── Sheet 4: Coverage Pivot ────────────────────────────────
        ws_cov = wb.create_sheet("Coverage")
        ws_cov.append(["Namespace", "Total", "Full", "Partial", "None", "Coverage %"])
        _style_header_row(ws_cov)
        for i, w in enumerate([16, 10, 10, 10, 10, 14], start=1):
            ws_cov.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w
        ns_stats: Dict[str, Dict[str, int]] = {}
        for ev in events:
            ns = ev["namespace"]
            assigned_agent = registry.get_event_agent(ev["event_id"])
            if not assigned_agent:
                claims = ns_to_agent_default.get(ns, [])
                if claims:
                    assigned_agent = claims[0]
            tool_info = None
            if assigned_agent and assigned_agent in agent_lookup:
                tool_info = discovery.find_tool_for_event(
                    ev["event_id"], agent_lookup[assigned_agent].get("mcp_servers", []),
                )
            cov = "full" if (assigned_agent and tool_info) else (
                "partial" if assigned_agent else "none"
            )
            ns_stats.setdefault(ns, {"total": 0, "full": 0, "partial": 0, "none": 0})
            ns_stats[ns]["total"] += 1
            ns_stats[ns][cov] += 1

        for ns in sorted(ns_stats.keys()):
            s = ns_stats[ns]
            pct = round(100 * s["full"] / max(1, s["total"]), 1)
            row_idx = ws_cov.max_row + 1
            ws_cov.append([ns, s["total"], s["full"], s["partial"], s["none"], pct])
            # Color full cell green, partial yellow, none red
            ws_cov.cell(row=row_idx, column=3).fill = FULL_FILL
            ws_cov.cell(row=row_idx, column=4).fill = PARTIAL_FILL
            ws_cov.cell(row=row_idx, column=5).fill = NONE_FILL
            for c in range(3, 6):
                ws_cov.cell(row=row_idx, column=c).font = Font(bold=True, color="FFFFFF")
                ws_cov.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center")
        # Totals row
        total_row = ws_cov.max_row + 1
        ws_cov.append(["TOTAL", len(events), full_count, partial_count, none_count,
                       round(100 * full_count / max(1, len(events)), 1)])
        for c in range(1, 7):
            cell = ws_cov.cell(row=total_row, column=c)
            cell.font = Font(bold=True, color="4FC3F7")
            cell.fill = HEADER_FILL

        # ─── Stream as binary ───────────────────────────────────────
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import datetime
        fname = f"vibemind_event_mapping_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as e:
        import traceback
        return JSONResponse(
            {"ok": False, "error": str(e), "trace": traceback.format_exc()[:1000]},
            status_code=500,
        )


@router.post("/api/ui/refresh-bubbles")
async def ui_refresh_bubbles(request: Request) -> JSONResponse:
    """Phase 11.U — force the UI to re-fetch bubbles from the source of
    truth. Useful after external mutations (direct DB wipe, supabase
    UI delete, manual rowboat manifest cleanup) where no per-bubble
    bubble.delete event was emitted.

    Publishes a single `ui.refresh_bubbles` space-event. The Electron
    brain-event-bridge maps this to an IPC `force_resync_bubbles`
    message which the renderer responds to by calling
    `vibemind.requestBubbles()`. The Voice subprocess then re-loads
    from Supabase (Phase 11.P force_reload=True) and ships back
    `bubbles_sync` to the renderer."""
    try:
        from core.space_event_bus import get_bus
        bus = get_bus()
        if bus._publish_loop is None:
            try:
                bus.attach_loop(asyncio.get_running_loop())
            except Exception:
                pass
        return JSONResponse(bus.publish({
            "event_id": "ui.refresh_bubbles",
            "params": {},
            "ok": True,
            "result": "force resync requested",
            "source": "api/ui/refresh-bubbles",
        }))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.post("/api/events/publish")
async def events_publish(request: Request) -> JSONResponse:
    """Phase 11.F — receive an event from a tool (in any process).
    Body: {event_id, params?, result?, ok?, source?, agent?, plan_id?, context?}
    """
    try:
        from core.space_event_bus import get_bus
        body = await request.json()
        bus = get_bus()
        if bus._publish_loop is None:
            try:
                bus.attach_loop(asyncio.get_running_loop())
            except Exception:
                pass
        return JSONResponse(bus.publish(body or {}))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/events/stream")
async def events_stream(request: Request):
    """Phase 11.F — SSE stream of all space-events (bubble.*, idea.*, etc).
    First message is the recent ring (last 30); subsequent are live."""
    from fastapi.responses import StreamingResponse
    from core.space_event_bus import get_bus

    bus = get_bus()
    try:
        bus.attach_loop(asyncio.get_running_loop())
    except Exception:
        pass
    q = bus.subscribe()

    async def gen():
        try:
            recent = bus.recent(30)
            yield f"event: recent\ndata: {json.dumps(recent)}\n\n"
            yield "event: ready\ndata: {}\n\n"
            while True:
                try:
                    ev = await asyncio.wait_for(q.get(), timeout=20.0)
                    yield f"event: space_event\ndata: {json.dumps(ev)}\n\n"
                except asyncio.TimeoutError:
                    yield ": keepalive\n\n"
        except asyncio.CancelledError:
            raise
        finally:
            bus.unsubscribe(q)

    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache",
                                      "X-Accel-Buffering": "no"})


@router.get("/api/events/recent")
async def events_recent(request: Request, limit: int = 50) -> JSONResponse:
    try:
        from core.space_event_bus import get_bus
        return JSONResponse({
            "ok": True,
            "events": get_bus().recent(limit=limit),
            "stats": get_bus().stats_dict(),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/mcp/discovery_stats")
async def mcp_discovery_stats(request: Request) -> JSONResponse:
    try:
        from core.mcp_discovery import get_discovery
        return JSONResponse({
            "ok": True,
            "stats": get_discovery().stats_dict(),
            "servers": get_discovery().list_servers(),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ===================================================================
# UI page
# ===================================================================

@router.get("/ui/brain", response_class=HTMLResponse)
async def brain_ui(request: Request) -> HTMLResponse:
    """Render brain dashboard."""
    return request.app.state.templates.TemplateResponse(
        request, "brain_dashboard.html"
    )


def _count_complete_lines(path, start_offset: int = 0) -> int:
    """Count COMPLETE (newline-terminated) lines in `path` from byte
    `start_offset` to EOF, without loading the file into memory — seek, then
    stream 1 MB chunks counting b'\\n'. A trailing partial line (no
    terminating '\\n', i.e. an in-flight write) is deliberately NOT counted:
    we tally newlines seen, not "lines" in the naive splitlines() sense, so a
    dangling fragment after the last '\\n' never contributes."""
    count = 0
    with open(path, "rb") as f:
        if start_offset > 0:
            f.seek(start_offset)
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            count += chunk.count(b"\n")
    return count


def _diary_queue_block() -> Dict[str, Any]:
    """Worker-independent queue/drain snapshot — reads the shared-volume
    FILES (queue + drain state), never in-memory state, so it reflects
    reality regardless of which brain-core worker process answers this
    request. Degrades to zeros on any failure; never raises.

    `pending` is derived from the drain's byte OFFSET, not from
    `enqueued - episodes_drained`. Those two counters do NOT reconcile: the
    drain advances the offset past corrupt / structurally-unusable /
    backstop-abandoned lines (they are consumed and gone) but never counts
    them as drained. Subtracting would therefore overstate the backlog by the
    number of permanently-skipped lines and never reach 0 even when the drain
    is fully caught up. The offset is the authoritative "how far consumed"
    marker, so the exact, skip-immune backlog is the number of complete lines
    in [offset, EOF) — which is also cheaper, since we only scan the
    un-drained tail.
    """
    try:
        from core.multihop_kotlin_adapter import resolve_queue_path
        from core.multihop_diary_drain import _default_state_path

        q_path = resolve_queue_path()
        if not q_path.exists():
            return {
                "episodes_enqueued": 0, "episodes_drained": 0, "skipped": 0,
                "pending": 0, "last_plan_id": None, "path": str(q_path),
            }

        enqueued = _count_complete_lines(q_path)

        # Missing OR corrupt state file -> offset 0: we cannot prove anything
        # was drained, so everything in the queue counts as pending. Honest
        # under-confidence beats a fabricated number.
        offset = 0
        drained = 0
        skipped = 0
        last_plan_id = None
        state_path = _default_state_path(q_path)
        if state_path.exists():
            try:
                state = json.loads(state_path.read_text(encoding="utf-8"))
                if isinstance(state, dict):
                    offset = int(state.get("offset", 0) or 0)
                    # NOTE: `episodes_drained` = episodes actually REPLAYED
                    # (incl. idempotent duplicates). It is informative, but it
                    # is NOT the complement of `pending` — do not re-derive
                    # pending from it (see the docstring). `skipped` is what
                    # makes the books balance:
                    #     enqueued ≈ drained + skipped + pending
                    drained = int(state.get("episodes_drained", 0) or 0)
                    skipped = int(state.get("lines_skipped", 0) or 0)
                    last_plan_id = state.get("last_plan_id") or None
            except Exception:
                pass

        size = q_path.stat().st_size
        if offset < 0 or offset > size:
            # Rotation/truncation: mirror the drain's own reset rule rather
            # than counting backwards into a negative pending.
            offset = 0
        pending = _count_complete_lines(q_path, start_offset=offset)

        return {
            "episodes_enqueued": enqueued,
            "episodes_drained": drained,
            "skipped": skipped,
            "pending": pending,
            "last_plan_id": last_plan_id,
            "path": str(q_path),
        }
    except Exception:
        return {
            "episodes_enqueued": 0,
            "episodes_drained": 0,
            "skipped": 0,
            "pending": 0,
            "last_plan_id": None,
            "path": "",
        }


def _diary_enqueue_block(request: Request) -> Dict[str, Any]:
    """plan_executor's diary_enqueued/diary_enqueue_failures counters — a
    failed enqueue is a silently dropped episode, so this is the last place
    to notice that data loss. Degrades to zeros on any failure."""
    try:
        pe = getattr(request.app.state, "plan_executor", None)
        if pe is None:
            return {"ok": 0, "failures": 0}
        stats = pe.stats_dict() or {}
        return {
            "ok": int(stats.get("diary_enqueued", 0) or 0),
            "failures": int(stats.get("diary_enqueue_failures", 0) or 0),
        }
    except Exception:
        return {"ok": 0, "failures": 0}


@router.get("/api/diary/stats")
async def diary_stats(request: Request):
    """Phase 1 — Read-only Blick ins episodische Tagebuch (KotlinGraph).
    Grundlage für den Live-Beweis: schreibt der Multihop-Ingest real?

    brain-core no longer writes episodes into its own in-memory dual_graph
    (2 uvicorn workers, never persists) — it appends to a shared queue that
    a separate drain process replays. So `multihop_events` etc. below stay
    honestly empty on brain-core; the `queue`/`enqueue` blocks are the
    worker-independent, file-based signal that actually shows episodes
    flowing."""
    dg = getattr(request.app.state, "dual_graph", None)
    if dg is None:
        return JSONResponse({"error": "dual_graph not loaded"}, status_code=503)
    try:
        kg = dg.kotlingraph
        multihop = sum(
            1 for e in kg.events
            if (getattr(e, "metadata", None) or {}).get("source") == "multihop"
        )
        last = kg.events[-1] if kg.events else None
        return JSONResponse({
            "total_events": kg.stats.get("total_events", 0),
            "total_episodes": kg.stats.get("total_episodes", 0),
            "multihop_events": multihop,
            "current_episode_id": kg.current_episode_id,
            "last_event": ({
                "action": last.action,
                "done": last.done,
                "reward": last.reward,
                "plan_id": (last.metadata or {}).get("plan_id"),
                "episode_success": (last.metadata or {}).get("episode_success"),
            } if last else None),
            "queue": _diary_queue_block(),
            "enqueue": _diary_enqueue_block(request),
        })
    except Exception as e:  # noqa: BLE001 — Introspection darf nie crashen
        return JSONResponse({"error": str(e)[:200]}, status_code=500)


@router.get("/api/toolscope")
async def toolscope_debug(intent: str, agent: str = "skill-coordinator", top_n: int = 8):
    """Debug/Verifikation der dynamischen Tool-Auswahl (plans/dynamic-agent-tools-prompt.md).

    Ruft den ToolScopeSelector mit dem WARMEN Prozess-Embedder (uvicorn hat ihn
    geladen) und gibt die gewaehlten Tools + Prompt-Focus sofort zurueck — umgeht
    den langsamen openfang-/execute-Roundtrip. Read-only, kein Seiteneffekt.
    Beantwortet: 'waehlt der Selektor live sinnvolle Tools fuer diesen Intent?'.
    """
    import time as _t
    t0 = _t.time()
    try:
        from core.tool_scope_selector import get_selector
        # Embedder im Threadpool ziehen, falls (auf diesem Worker) noch kalt —
        # blockiert dann nicht den Event-Loop.
        import asyncio as _a
        loop = _a.get_running_loop()
        allow, focus = await loop.run_in_executor(
            None, lambda: get_selector().select_tools(intent, agent_name=agent, top_n=top_n))
        return JSONResponse({
            "intent": intent, "agent": agent, "top_n": top_n,
            "tool_count": len(allow), "tools": allow,
            "prompt_focus": focus,
            "elapsed_s": round(_t.time() - t0, 2),
        })
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({
            "intent": intent, "agent": agent, "error": str(exc),
            "elapsed_s": round(_t.time() - t0, 2),
        }, status_code=500)
